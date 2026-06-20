import customtkinter as ctk
from tkinter import ttk, messagebox, filedialog
import tkinter as tk
from datetime import date, datetime, timedelta
from app.database.connection import get_db_session
from app.database.models import Venta, ItemVenta, Producto, EstadoVenta, Lote, RolUsuario, MetodoPago, CortesCaja
from sqlalchemy import func
import app.config as cfg

def _get_admin_pin() -> str:
    """Lee el PIN admin desde DB; si no existe usa el valor por defecto."""
    try:
        from app.database.connection import get_db_session
        from app.database.models import Configuracion
        db = get_db_session()
        try:
            row = db.query(Configuracion).filter(Configuracion.clave == "admin_pin").first()
            return row.valor if row and row.valor else "171215"
        finally:
            db.close()
    except Exception:
        return "171215"

# Palette (matches main_window)
CARD_BG = "#FFFFFF"
CONT_BG = "#F0F4F8"
BORDER  = "#E2E8F0"
TEXT    = "#0F172A"
MUTED   = "#64748B"
BLUE    = "#2563EB"
BLUE_L  = "#EFF6FF"
GREEN   = "#22C55E"
GREEN_L = "#F0FDF4"
AMBER   = "#F59E0B"
AMBER_L = "#FFFBEB"
PURPLE  = "#8B5CF6"
PURPLE_L= "#F5F3FF"

_CARD_DEFS = [
    ("total_ventas", "Total Ventas",      "$0.00",  BLUE,   BLUE_L,   "💰"),
    ("num_ventas",   "N° Transacciones",  "0",      GREEN,  GREEN_L,  "🛒"),
    ("ticket_prom",  "Ticket Promedio",   "$0.00",  AMBER,  AMBER_L,  "📈"),
    ("top_dia",      "Mejor Día",         "$0.00",  PURPLE, PURPLE_L, "🏆"),
]

_PAGO_DEFS = [
    ("efectivo",       "Efectivo",       GREEN,  GREEN_L,  "💵"),
    ("tarjeta",        "Tarjeta",        BLUE,   BLUE_L,   "💳"),
    ("transferencia",  "Transferencia",  PURPLE, PURPLE_L, "🏦"),
]


class ReportsScreen(ctk.CTkFrame):
    def __init__(self, parent, user):
        super().__init__(parent, corner_radius=0, fg_color="transparent")
        self.user = user
        self._build_ui()

    # ── UI shell ──────────────────────────────────────────────────────────────

    def _build_ui(self):
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(1, weight=1)

        self._build_filter_bar()
        self._build_tabs()
        self._generar()

    def _build_filter_bar(self):
        bar = ctk.CTkFrame(self, corner_radius=12, fg_color=CARD_BG,
                           border_width=1, border_color=BORDER)
        bar.grid(row=0, column=0, sticky="ew", padx=14, pady=(14, 6))
        bar.grid_columnconfigure(1, weight=1)

        # Title
        ctk.CTkLabel(bar, text="Reportes",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=TEXT).grid(row=0, column=0, padx=(18, 20), pady=14, sticky="w")

        # Date filters
        ff = ctk.CTkFrame(bar, fg_color="transparent")
        ff.grid(row=0, column=1, sticky="w", pady=10)

        ctk.CTkLabel(ff, text="Desde",
                     font=ctk.CTkFont(size=11), text_color=MUTED).pack(side="left", padx=(0, 4))
        self.entry_desde = ctk.CTkEntry(ff, width=106, height=34,
                                        placeholder_text="DD/MM/YYYY",
                                        corner_radius=8, border_color=BORDER,
                                        fg_color="#F8FAFF", font=ctk.CTkFont(size=12))
        self.entry_desde.pack(side="left", padx=(0, 14))
        hoy = date.today()
        inicio_mes = date(hoy.year, hoy.month, 1)
        self.entry_desde.insert(0, inicio_mes.strftime("%d/%m/%Y"))

        ctk.CTkLabel(ff, text="Hasta",
                     font=ctk.CTkFont(size=11), text_color=MUTED).pack(side="left", padx=(0, 4))
        self.entry_hasta = ctk.CTkEntry(ff, width=106, height=34,
                                         placeholder_text="DD/MM/YYYY",
                                         corner_radius=8, border_color=BORDER,
                                         fg_color="#F8FAFF", font=ctk.CTkFont(size=12))
        self.entry_hasta.pack(side="left", padx=(0, 18))
        self.entry_hasta.insert(0, hoy.strftime("%d/%m/%Y"))

        # Quick-period chips
        chips = ctk.CTkFrame(ff, fg_color="transparent")
        chips.pack(side="left")
        for lbl, days in [("Hoy", 0), ("7 días", 7), ("Este mes", -1), ("Mes ant.", -2), ("Todas", -99)]:
            ctk.CTkButton(
                chips, text=lbl, width=74, height=30, corner_radius=8,
                fg_color=BLUE_L, hover_color="#DBEAFE", text_color=BLUE,
                font=ctk.CTkFont(size=11, weight="bold"),
                command=lambda d=days: self._set_periodo(d),
            ).pack(side="left", padx=2)

        # Action buttons
        ab = ctk.CTkFrame(bar, fg_color="transparent")
        ab.grid(row=0, column=2, padx=(0, 14), pady=10, sticky="e")
        ctk.CTkButton(
            ab, text="Generar", width=90, height=34, corner_radius=8,
            fg_color=BLUE, hover_color="#1D4ED8",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._generar,
        ).pack(side="left", padx=(0, 6))
        ctk.CTkButton(
            ab, text="Excel", width=80, height=34, corner_radius=8,
            fg_color=GREEN_L, hover_color="#DCFCE7", text_color=GREEN,
            border_width=1, border_color="#BBF7D0",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=self._exportar_excel,
        ).pack(side="left", padx=(0, 6))

        if self.user.rol == RolUsuario.admin:
            ctk.CTkButton(
                ab, text="🗑  Purgar datos", width=120, height=34, corner_radius=8,
                fg_color="#FEF2F2", hover_color="#FEE2E2", text_color="#EF4444",
                border_width=1, border_color="#FECACA",
                font=ctk.CTkFont(size=12, weight="bold"),
                command=self._dlg_purgar,
            ).pack(side="left")

    def _build_tabs(self):
        self.tabs = ctk.CTkTabview(self, fg_color=CONT_BG,
                                   segmented_button_fg_color=CARD_BG,
                                   segmented_button_selected_color=BLUE,
                                   segmented_button_selected_hover_color="#1D4ED8",
                                   segmented_button_unselected_color=CARD_BG,
                                   segmented_button_unselected_hover_color=BORDER,
                                   text_color=MUTED,
                                   text_color_disabled=MUTED)
        self.tabs.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        self.tabs.add("Resumen")
        self.tabs.add("Ventas Detalle")
        self.tabs.add("Productos Top")
        self.tabs.add("Cortes de Caja")
        self.tabs.add("Vencimientos")

        self._build_resumen_tab()
        self._build_ventas_tab()
        self._build_productos_tab()
        self._build_cortes_tab()
        self._build_vencimientos_tab()

    # ── Resumen tab ───────────────────────────────────────────────────────────

    def _build_resumen_tab(self):
        tab = self.tabs.tab("Resumen")
        tab.grid_columnconfigure((0, 1, 2, 3), weight=1)
        tab.grid_rowconfigure(1, weight=1)

        self.cards = {}
        for i, (key, title, default, accent, bg, icon) in enumerate(_CARD_DEFS):
            card = ctk.CTkFrame(tab, corner_radius=14, fg_color=CARD_BG,
                                border_width=1, border_color=BORDER)
            card.grid(row=0, column=i, padx=(0 if i else 0, 8 if i < 3 else 0),
                      pady=(0, 10), sticky="ew")
            card.grid_columnconfigure(0, weight=1)

            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=16, pady=16)

            # Left: texts
            tf = ctk.CTkFrame(inner, fg_color="transparent")
            tf.pack(side="left", fill="both", expand=True)

            lbl = ctk.CTkLabel(tf, text=default,
                               font=ctk.CTkFont(size=22, weight="bold"),
                               text_color=accent, anchor="w")
            lbl.pack(anchor="w")
            ctk.CTkLabel(tf, text=title,
                         font=ctk.CTkFont(size=10),
                         text_color=MUTED, anchor="w").pack(anchor="w")

            # Right: icon circle
            ic = ctk.CTkFrame(inner, width=42, height=42,
                              corner_radius=21, fg_color=bg)
            ic.pack(side="right")
            ic.pack_propagate(False)
            ctk.CTkLabel(ic, text=icon,
                         font=ctk.CTkFont(size=17)).place(relx=0.5, rely=0.5, anchor="center")

            self.cards[key] = lbl

        # Payment methods row
        pf = ctk.CTkFrame(tab, corner_radius=14, fg_color=CARD_BG,
                          border_width=1, border_color=BORDER)
        pf.grid(row=1, column=0, columnspan=4, sticky="nsew")

        ctk.CTkLabel(pf, text="Ventas por Método de Pago",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color=TEXT).pack(anchor="w", padx=18, pady=(16, 10))

        ctk.CTkFrame(pf, height=1, fg_color=BORDER).pack(fill="x", padx=18, pady=(0, 12))

        self.pago_labels = {}
        pr = ctk.CTkFrame(pf, fg_color="transparent")
        pr.pack(fill="x", padx=14, pady=(0, 16))
        pr.grid_columnconfigure((0, 1, 2), weight=1)

        for i, (metodo, label, accent, bg, icon) in enumerate(_PAGO_DEFS):
            pc = ctk.CTkFrame(pr, corner_radius=12, fg_color=bg,
                              border_width=1, border_color=BORDER)
            pc.grid(row=0, column=i, padx=(0 if i else 0, 8 if i < 2 else 0), sticky="ew")

            pi = ctk.CTkFrame(pc, fg_color="transparent")
            pi.pack(fill="x", padx=16, pady=14)

            ic = ctk.CTkFrame(pi, width=36, height=36,
                              corner_radius=18, fg_color=CARD_BG)
            ic.pack(side="left", padx=(0, 10))
            ic.pack_propagate(False)
            ctk.CTkLabel(ic, text=icon,
                         font=ctk.CTkFont(size=14)).place(relx=0.5, rely=0.5, anchor="center")

            tt = ctk.CTkFrame(pi, fg_color="transparent")
            tt.pack(side="left")
            ctk.CTkLabel(tt, text=label,
                         font=ctk.CTkFont(size=10), text_color=MUTED, anchor="w").pack(anchor="w")
            val = ctk.CTkLabel(tt, text="$0.00",
                               font=ctk.CTkFont(size=15, weight="bold"),
                               text_color=accent, anchor="w")
            val.pack(anchor="w")
            self.pago_labels[metodo] = val

    # ── Detail tabs ───────────────────────────────────────────────────────────

    def _make_tree_style(self, name: str) -> str:
        style = ttk.Style()
        style.configure(f"{name}.Treeview",
                        background="#FFFFFF", foreground=TEXT,
                        rowheight=30, fieldbackground="#FFFFFF",
                        borderwidth=0, font=("Segoe UI", 10))
        style.configure(f"{name}.Treeview.Heading",
                        background="#F1F5F9", foreground=MUTED,
                        relief="flat", font=("Segoe UI", 10, "bold"),
                        padding=(8, 6))
        style.map(f"{name}.Treeview",
                  background=[("selected", "#DBEAFE")],
                  foreground=[("selected", BLUE)])
        style.layout(f"{name}.Treeview", [
            ("Treeview.treearea", {"sticky": "nswe"})
        ])
        return f"{name}.Treeview"

    def _build_ventas_tab(self):
        tab = self.tabs.tab("Ventas Detalle")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        self._venta_ids: dict = {}  # tree iid → venta.id

        # Admin toolbar
        if self.user.rol == RolUsuario.admin:
            tb = ctk.CTkFrame(tab, fg_color="transparent")
            tb.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 4))
            ctk.CTkButton(
                tb,
                text="🗑  Eliminar seleccionada",
                width=190, height=28, corner_radius=8,
                fg_color="#FEF2F2", hover_color="#FEE2E2", text_color="#EF4444",
                border_width=1, border_color="#FECACA",
                font=ctk.CTkFont(size=11, weight="bold"),
                command=self._eliminar_venta_seleccionada,
            ).pack(side="right", padx=4)

        sty = self._make_tree_style("V")
        cols = ("folio", "fecha", "cajero", "cliente", "subtotal", "iva", "total", "pago", "estado")
        self.ventas_tree = ttk.Treeview(tab, columns=cols, show="headings", style=sty)
        for col, heading, width in [
            ("folio",    "Folio",      110), ("fecha",    "Fecha/Hora", 130),
            ("cajero",   "Cajero",     100), ("cliente",  "Cliente",    120),
            ("subtotal", "Subtotal",    80), ("iva",      "IVA",         65),
            ("total",    "Total",       80), ("pago",     "Pago",        90),
            ("estado",   "Estado",      90),
        ]:
            self.ventas_tree.heading(col, text=heading)
            self.ventas_tree.column(col, width=width)

        # Alternating row colors
        self.ventas_tree.tag_configure("even", background="#F8FAFF")
        self.ventas_tree.tag_configure("odd",  background="#FFFFFF")

        scroll = ttk.Scrollbar(tab, orient="vertical", command=self.ventas_tree.yview)
        self.ventas_tree.configure(yscrollcommand=scroll.set)
        self.ventas_tree.grid(row=1, column=0, sticky="nsew")
        scroll.grid(row=1, column=1, sticky="ns")

        self.ventas_tree.bind("<Double-1>", self._ver_detalle_venta)
        if self.user.rol == RolUsuario.admin:
            self.ventas_tree.bind("<Button-3>", self._ctx_menu_venta)

    def _build_productos_tab(self):
        tab = self.tabs.tab("Productos Top")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(0, weight=1)

        sty = self._make_tree_style("P")
        cols = ("pos", "producto", "cantidad", "ingresos")
        self.prod_tree = ttk.Treeview(tab, columns=cols, show="headings", style=sty)
        for col, heading, width in [
            ("pos",      "#",                  40),
            ("producto", "Producto",          300),
            ("cantidad", "Unidades Vendidas", 140),
            ("ingresos", "Ingresos",          100),
        ]:
            self.prod_tree.heading(col, text=heading)
            self.prod_tree.column(col, width=width)

        self.prod_tree.tag_configure("even", background="#F8FAFF")
        self.prod_tree.tag_configure("odd",  background="#FFFFFF")

        scroll = ttk.Scrollbar(tab, orient="vertical", command=self.prod_tree.yview)
        self.prod_tree.configure(yscrollcommand=scroll.set)
        self.prod_tree.grid(row=0, column=0, sticky="nsew")
        scroll.grid(row=0, column=1, sticky="ns")

    def _build_cortes_tab(self):
        tab = self.tabs.tab("Cortes de Caja")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(2, weight=1)

        # ── Summary cards ─────────────────────────────────────────────────────
        cf = ctk.CTkFrame(tab, corner_radius=10, fg_color=CARD_BG,
                          border_width=1, border_color=BORDER)
        cf.grid(row=0, column=0, columnspan=2, sticky="ew", pady=(0, 8))
        cf.grid_columnconfigure((0, 1, 2, 3, 4), weight=1)

        self._corte_cards = {}
        for i, (key, label, color, bg) in enumerate([
            ("c_turnos",        "Turnos",          PURPLE, PURPLE_L),
            ("c_total_ventas",  "Total Ventas",    BLUE,   BLUE_L),
            ("c_efectivo",      "Efectivo",        GREEN,  GREEN_L),
            ("c_tarjeta",       "Tarjeta/Transf.", AMBER,  AMBER_L),
            ("c_diferencia",    "Diferencia",      "#EF4444", "#FEF2F2"),
        ]):
            card = ctk.CTkFrame(cf, corner_radius=8, fg_color=bg)
            card.grid(row=0, column=i, padx=(14 if i == 0 else 5, 14 if i == 4 else 5),
                      pady=10, sticky="ew")
            ctk.CTkLabel(card, text=label,
                         font=ctk.CTkFont(size=9, weight="bold"),
                         text_color=MUTED).pack(pady=(8, 0))
            val = ctk.CTkLabel(card, text="—",
                               font=ctk.CTkFont(size=15, weight="bold"),
                               text_color=color)
            val.pack(pady=(2, 8))
            self._corte_cards[key] = val

        # ── Treeview ──────────────────────────────────────────────────────────
        sty = self._make_tree_style("C")
        cols = ("cajero", "apertura_dt", "cierre_dt", "duracion",
                "num_ventas", "total_ventas", "efectivo", "tarjeta",
                "transferencia", "fondo_apertura", "fondo_cierre",
                "diferencia", "estado")
        self.cortes_tree = ttk.Treeview(tab, columns=cols, show="headings", style=sty)
        for col, heading, width in [
            ("cajero",          "Cajero",           140),
            ("apertura_dt",     "Apertura",         130),
            ("cierre_dt",       "Cierre",           130),
            ("duracion",        "Duración",          75),
            ("num_ventas",      "# Ventas",          65),
            ("total_ventas",    "Total Ventas",      100),
            ("efectivo",        "Efectivo",           90),
            ("tarjeta",         "Tarjeta",            90),
            ("transferencia",   "Transf.",            80),
            ("fondo_apertura",  "Fondo Apertura",    110),
            ("fondo_cierre",    "Fondo Cierre",      100),
            ("diferencia",      "Diferencia",         90),
            ("estado",          "Estado",             95),
        ]:
            self.cortes_tree.heading(col, text=heading)
            anchor = "w" if col == "cajero" else "center"
            self.cortes_tree.column(col, width=width, anchor=anchor)

        self.cortes_tree.tag_configure("even",      background="#F8FAFF")
        self.cortes_tree.tag_configure("odd",       background="#FFFFFF")
        self.cortes_tree.tag_configure("abierto",   foreground="#D97706", background="#FFFBEB")
        self.cortes_tree.tag_configure("descuadre", foreground="#DC2626", background="#FEF2F2")
        self.cortes_tree.tag_configure("cuadrado",  foreground="#16A34A")

        scroll_y = ttk.Scrollbar(tab, orient="vertical",   command=self.cortes_tree.yview)
        scroll_x = ttk.Scrollbar(tab, orient="horizontal", command=self.cortes_tree.xview)
        self.cortes_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.cortes_tree.grid(row=2, column=0, sticky="nsew")
        scroll_y.grid(row=2, column=1, sticky="ns")
        scroll_x.grid(row=3, column=0, columnspan=2, sticky="ew")

        self.cortes_tree.bind("<Double-1>", self._ver_detalle_corte)

        # ── Botón cierre rápido (solo admin, visible solo cuando hay turno abierto) ──
        if self.user.rol == RolUsuario.admin:
            self._btn_cerrar_activo = ctk.CTkButton(
                tab,
                text="🔒 Cerrar turno abierto ahora",
                height=38, corner_radius=8,
                fg_color="#F59E0B", hover_color="#D97706", text_color="white",
                font=ctk.CTkFont(size=12, weight="bold"),
                command=self._cerrar_turno_activo_rapido,
            )
            self._btn_cerrar_activo.grid(row=1, column=0, columnspan=2,
                                          sticky="w", padx=4, pady=(0, 6))
            self._btn_cerrar_activo.grid_remove()  # hidden until open turno detected

    def _cerrar_turno_activo_rapido(self):
        """Cierra todos los turnos abiertos directamente desde el botón en la pestaña."""
        from app.database.connection import get_db_session
        from app.database.models import CortesCaja, Venta, EstadoVenta, MetodoPago, ItemVenta, Producto
        from datetime import datetime as _dt
        db = get_db_session()
        try:
            cortes = db.query(CortesCaja).filter(CortesCaja.cerrado_en == None).all()
            if not cortes:
                messagebox.showinfo("Turnos", "No hay turnos abiertos.")
                return
            nombres = ", ".join(c.usuario.nombre if c.usuario else f"#{c.id}" for c in cortes)
            if not messagebox.askyesno(
                "Cerrar turno",
                f"¿Cerrar el turno activo de:\n{nombres}?\n\nEsta acción no se puede deshacer."
            ):
                return
            ahora = _dt.now()
            for c in cortes:
                vq = db.query(Venta).filter(
                    Venta.usuario_id == c.usuario_id,
                    Venta.creado_en >= c.abierto_en,
                    Venta.creado_en <= ahora,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True),
                ).all()
                c_ef = sum(v.total for v in vq if v.metodo_pago == MetodoPago.efectivo)
                c_tj = sum(v.total for v in vq if v.metodo_pago == MetodoPago.tarjeta)
                c_tr = sum(v.total for v in vq if v.metodo_pago == MetodoPago.transferencia)
                c_tv = c_ef + c_tj + c_tr
                vids = [v.id for v in vq]
                if vids:
                    cost_rows = (
                        db.query(ItemVenta.cantidad, Producto.precio_compra)
                        .join(Producto, ItemVenta.producto_id == Producto.id)
                        .filter(ItemVenta.venta_id.in_(vids))
                        .all()
                    )
                    tc = sum(r.cantidad * (r.precio_compra or 0.0) for r in cost_rows)
                else:
                    tc = 0.0
                c.cerrado_en          = ahora
                c.total_ventas        = c_tv
                c.total_efectivo      = c_ef
                c.total_tarjeta       = c_tj
                c.total_transferencia = c_tr
                c.total_costo         = tc
                c.num_ventas          = len(vq)
                c.monto_cierre        = c.monto_apertura + c_ef
                notas_prev            = (c.notas or "").strip()
                c.notas               = (notas_prev + " [Cierre admin desde reportes]").strip()
            db.commit()
            messagebox.showinfo(
                "Turno cerrado",
                f"Turno(s) cerrado(s) correctamente.\nCajero(s): {nombres}"
            )
            self._generar()
        except Exception as exc:
            db.rollback()
            messagebox.showerror("Error", str(exc))
        finally:
            db.close()

    def _ver_detalle_corte(self, event=None):
        sel = self.cortes_tree.selection()
        if not sel:
            return
        corte_id = int(sel[0])   # iid = str(c.id)
        vals = self.cortes_tree.item(sel[0], "values")
        if not vals:
            return
        # vals: cajero, apertura_dt, cierre_dt, duracion, num_ventas, total_ventas,
        #       efectivo, tarjeta, transferencia, fondo_apertura, fondo_cierre, diferencia, estado
        cajero, ap_dt, ci_dt, dur, nv, tv, ef, tj, tr, fa, fc, dif, estado = vals

        win = ctk.CTkToplevel(self)
        win.title("Detalle de Corte de Caja")
        win.geometry("480x520")
        win.grab_set()

        ctk.CTkLabel(win, text="🧾 Detalle de Corte",
                     font=ctk.CTkFont(size=16, weight="bold")).pack(pady=(18, 4))
        ctk.CTkLabel(win, text=f"Cajero: {cajero}",
                     font=ctk.CTkFont(size=13), text_color=MUTED).pack()

        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=12)

        def row(parent, label, value, color=TEXT):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", padx=24, pady=3)
            ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=12),
                         text_color=MUTED, anchor="w", width=200).pack(side="left")
            ctk.CTkLabel(f, text=value, font=ctk.CTkFont(size=12, weight="bold"),
                         text_color=color, anchor="e").pack(side="right")

        ctk.CTkLabel(win, text="Horario", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=MUTED, anchor="w").pack(fill="x", padx=24, pady=(0, 2))
        row(win, "Apertura", ap_dt)
        row(win, "Cierre",   ci_dt)
        row(win, "Duración", dur)

        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=10)
        ctk.CTkLabel(win, text="Ventas del Turno", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=MUTED, anchor="w").pack(fill="x", padx=24, pady=(0, 2))
        row(win, "# Transacciones", nv)
        row(win, "Total Ventas",    tv,  BLUE)
        row(win, "Efectivo",        ef,  GREEN)
        row(win, "Tarjeta",         tj,  AMBER)
        row(win, "Transferencia",   tr,  PURPLE)

        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=10)
        ctk.CTkLabel(win, text="Cuadre de Efectivo", font=ctk.CTkFont(size=11, weight="bold"),
                     text_color=MUTED, anchor="w").pack(fill="x", padx=24, pady=(0, 2))
        row(win, "Fondo Apertura",          fa)
        row(win, "Efectivo ingresado (ventas)", ef, GREEN)

        try:
            fa_v = float(fa.replace("$", "").replace(",", ""))
            ef_v = float(ef.replace("$", "").replace(",", ""))
            esperado = fa_v + ef_v
            row(win, "Total esperado en caja", f"${esperado:.2f}", TEXT)
        except Exception:
            row(win, "Total esperado en caja", "—", MUTED)

        row(win, "Fondo Cierre (contado)",  fc)
        dif_color = "#EF4444" if dif.startswith("-") or (dif not in ("—", "$0.00", "$+0.00")) else "#16A34A"
        row(win, "Diferencia", dif, dif_color)

        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=10)
        estado_color = "#D97706" if "Abierto" in estado else (
            "#DC2626" if "Descuadre" in estado else "#16A34A")
        ctk.CTkLabel(win, text=estado,
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color=estado_color).pack(pady=(0, 8))

        # Botón de cierre manual (solo admin, solo turnos abiertos)
        if "Abierto" in estado and self.user.rol == RolUsuario.admin:
            def _cerrar_turno():
                from app.database.connection import get_db_session
                from app.database.models import CortesCaja, Venta, EstadoVenta, MetodoPago, ItemVenta, Producto
                from datetime import datetime as _dt
                db = get_db_session()
                try:
                    c = db.query(CortesCaja).filter(CortesCaja.id == corte_id).first()
                    if not c or c.cerrado_en:
                        messagebox.showinfo("", "El turno ya fue cerrado.")
                        win.destroy()
                        self._generar()
                        return
                    ahora = _dt.now()
                    vq = db.query(Venta).filter(
                        Venta.usuario_id == c.usuario_id,
                        Venta.creado_en >= c.abierto_en,
                        Venta.creado_en <= ahora,
                        Venta.estado == EstadoVenta.completada,
                        Venta.eliminado.is_not(True),
                    ).all()
                    c_ef = sum(v.total for v in vq if v.metodo_pago == MetodoPago.efectivo)
                    c_tj = sum(v.total for v in vq if v.metodo_pago == MetodoPago.tarjeta)
                    c_tr = sum(v.total for v in vq if v.metodo_pago == MetodoPago.transferencia)
                    c_tv = c_ef + c_tj + c_tr
                    vids = [v.id for v in vq]
                    if vids:
                        cost_rows = (
                            db.query(ItemVenta.cantidad, Producto.precio_compra)
                            .join(Producto, ItemVenta.producto_id == Producto.id)
                            .filter(ItemVenta.venta_id.in_(vids))
                            .all()
                        )
                        tc = sum(r.cantidad * (r.precio_compra or 0.0) for r in cost_rows)
                    else:
                        tc = 0.0
                    c.cerrado_en          = ahora
                    c.total_ventas        = c_tv
                    c.total_efectivo      = c_ef
                    c.total_tarjeta       = c_tj
                    c.total_transferencia = c_tr
                    c.total_costo         = tc
                    c.num_ventas          = len(vq)
                    c.monto_cierre        = c.monto_apertura + c_ef
                    notas_prev            = (c.notas or "").strip()
                    c.notas               = (notas_prev + " [Cierre manual desde reportes]").strip()
                    db.commit()
                    messagebox.showinfo("Turno cerrado",
                                        f"Turno cerrado correctamente.\nVentas: {len(vq)}  |  Total: ${c_tv:.2f}")
                    win.destroy()
                    self._generar()
                except Exception as exc:
                    db.rollback()
                    messagebox.showerror("Error", str(exc))
                finally:
                    db.close()

            ctk.CTkButton(
                win, text="🔒 Cerrar este turno ahora",
                height=38, fg_color="#F59E0B", hover_color="#D97706",
                text_color="white", font=ctk.CTkFont(size=13, weight="bold"),
                corner_radius=8,
                command=_cerrar_turno,
            ).pack(fill="x", padx=24, pady=(0, 16))

    # ── Periodo helpers ───────────────────────────────────────────────────────

    def _set_periodo(self, days: int):
        hoy = date.today()
        if days == -99:
            # Todas: fetch from 2020-01-01 to today
            desde = date(2020, 1, 1)
            hasta = hoy
            self.entry_desde.delete(0, "end")
            self.entry_desde.insert(0, desde.strftime("%d/%m/%Y"))
            self.entry_hasta.delete(0, "end")
            self.entry_hasta.insert(0, hasta.strftime("%d/%m/%Y"))
            self._generar()
            return
        if days == 0:
            desde = hasta = hoy
        elif days == -1:
            desde = date(hoy.year, hoy.month, 1)
            hasta = hoy
        elif days == -2:
            if hoy.month == 1:
                desde = date(hoy.year - 1, 12, 1)
                hasta = date(hoy.year - 1, 12, 31)
            else:
                desde = date(hoy.year, hoy.month - 1, 1)
                hasta = date(hoy.year, hoy.month, 1) - timedelta(days=1)
        else:
            desde = hoy - timedelta(days=days)
            hasta = hoy

        self.entry_desde.delete(0, "end")
        self.entry_desde.insert(0, desde.strftime("%d/%m/%Y"))
        self.entry_hasta.delete(0, "end")
        self.entry_hasta.insert(0, hasta.strftime("%d/%m/%Y"))
        self._generar()

    def _get_fechas(self):
        try:
            desde = datetime.strptime(self.entry_desde.get().strip(), "%d/%m/%Y")
            hasta = datetime.strptime(self.entry_hasta.get().strip(), "%d/%m/%Y").replace(
                hour=23, minute=59, second=59, microsecond=999999)
            return desde, hasta
        except ValueError:
            messagebox.showwarning("Fecha inválida", "Formato: DD/MM/YYYY")
            return None, None

    # ── Data logic (unchanged) ────────────────────────────────────────────────

    def _generar(self):
        desde, hasta = self._get_fechas()
        if not desde:
            return

        db = get_db_session()
        try:
            ventas = db.query(Venta).filter(
                Venta.creado_en >= desde,
                Venta.creado_en <= hasta,
                Venta.eliminado.is_not(True),
            ).order_by(Venta.creado_en.desc()).all()

            completadas = [v for v in ventas if v.estado == EstadoVenta.completada]
            total  = sum(v.total for v in completadas)
            num    = len(completadas)
            promedio = total / num if num > 0 else 0

            self.cards["total_ventas"].configure(text=f"${total:.2f}")
            self.cards["num_ventas"].configure(text=str(num))
            self.cards["ticket_prom"].configure(text=f"${promedio:.2f}")

            por_dia: dict[str, float] = {}
            for v in completadas:
                d = v.creado_en.strftime("%d/%m/%Y")
                por_dia[d] = por_dia.get(d, 0) + v.total
            mejor_dia = max(por_dia.values()) if por_dia else 0
            self.cards["top_dia"].configure(text=f"${mejor_dia:.2f}")

            for metodo in ["efectivo", "tarjeta", "transferencia"]:
                t = sum(v.total for v in completadas if v.metodo_pago.value == metodo)
                self.pago_labels[metodo].configure(text=f"${t:.2f}")

            for row in self.ventas_tree.get_children():
                self.ventas_tree.delete(row)
            self._venta_ids.clear()
            for idx, v in enumerate(ventas):
                tag = "even" if idx % 2 == 0 else "odd"
                iid = self.ventas_tree.insert("", "end", tags=(tag,), values=(
                    v.folio or v.id,
                    v.creado_en.strftime("%d/%m/%Y %H:%M") if v.creado_en else "",
                    v.usuario.nombre if v.usuario else "",
                    v.cliente.nombre if v.cliente else "Público",
                    f"${v.subtotal:.2f}", f"${v.iva:.2f}", f"${v.total:.2f}",
                    v.metodo_pago.value.capitalize(), v.estado.value.capitalize(),
                ))
                self._venta_ids[iid] = v.id

            conteo: dict = {}
            for v in completadas:
                for item in v.items:
                    pid = item.producto_id
                    nombre = item.producto.nombre if item.producto else "?"
                    if pid not in conteo:
                        conteo[pid] = {"nombre": nombre, "cantidad": 0, "ingresos": 0}
                    conteo[pid]["cantidad"] += item.cantidad
                    conteo[pid]["ingresos"] += item.subtotal

            for row in self.prod_tree.get_children():
                self.prod_tree.delete(row)
            ranking = sorted(conteo.values(), key=lambda x: x["cantidad"], reverse=True)[:20]
            for i, p in enumerate(ranking, 1):
                tag = "even" if i % 2 == 0 else "odd"
                self.prod_tree.insert("", "end", tags=(tag,), values=(
                    i, p["nombre"], p["cantidad"], f"${p['ingresos']:.2f}"
                ))

            cortes = db.query(CortesCaja).filter(
                CortesCaja.abierto_en >= desde,
                CortesCaja.abierto_en <= hasta,
            ).order_by(CortesCaja.abierto_en.desc()).all()

            for row in self.cortes_tree.get_children():
                self.cortes_tree.delete(row)

            # Recalculate each corte from actual ventas (fixes old stored zeros)
            corte_calc = []  # (c, ef, tj, tr, tv, num_v)
            for c in cortes:
                hasta_c = c.cerrado_en or datetime.now()
                vq = db.query(Venta).filter(
                    Venta.creado_en >= c.abierto_en,
                    Venta.creado_en <= hasta_c,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True),
                ).all()
                c_ef = sum(v.total for v in vq if v.metodo_pago == MetodoPago.efectivo)
                c_tj = sum(v.total for v in vq if v.metodo_pago == MetodoPago.tarjeta)
                c_tr = sum(v.total for v in vq if v.metodo_pago == MetodoPago.transferencia)
                c_tv = c_ef + c_tj + c_tr
                corte_calc.append((c, c_ef, c_tj, c_tr, c_tv, len(vq)))

            # Summary cards from recalculated values
            tot_v   = sum(tv         for _, _, _, _, tv, _  in corte_calc)
            tot_ef  = sum(ef         for _, ef, _, _, _, _  in corte_calc)
            tot_tj  = sum(tj + tr    for _, _, tj, tr, _, _ in corte_calc)
            tot_dif = sum(
                (c.monto_cierre or 0) - (c.monto_apertura + ef)
                for c, ef, _, _, _, _ in corte_calc if c.cerrado_en
            )
            hay_abierto = any(c.cerrado_en is None for c in cortes)
            if hasattr(self, "_btn_cerrar_activo"):
                if hay_abierto:
                    self._btn_cerrar_activo.grid()
                else:
                    self._btn_cerrar_activo.grid_remove()

            self._corte_cards["c_turnos"].configure(text=str(len(cortes)))
            self._corte_cards["c_total_ventas"].configure(text=f"${tot_v:.2f}")
            self._corte_cards["c_efectivo"].configure(text=f"${tot_ef:.2f}")
            self._corte_cards["c_tarjeta"].configure(text=f"${tot_tj:.2f}")
            dif_color = "#EF4444" if abs(tot_dif) > 0.01 else "#16A34A"
            self._corte_cards["c_diferencia"].configure(
                text=f"${tot_dif:+.2f}", text_color=dif_color)

            for c, c_ef, c_tj, c_tr, c_tv, num_v in corte_calc:
                dur_str = "—"
                if c.cerrado_en and c.abierto_en:
                    mins = int((c.cerrado_en - c.abierto_en).total_seconds() / 60)
                    dur_str = f"{mins // 60}h {mins % 60:02d}m"

                if c.cerrado_en:
                    dif = (c.monto_cierre or 0) - (c.monto_apertura + c_ef)
                    dif_str = f"${dif:+.2f}"
                    tag     = "descuadre" if abs(dif) > 0.01 else "cuadrado"
                    estado  = "⛔ Descuadre" if abs(dif) > 0.01 else "✓ Cuadrado"
                else:
                    dif_str = "—"
                    tag     = "abierto"
                    estado  = "🟡 Abierto"

                self.cortes_tree.insert("", "end", iid=str(c.id), tags=(tag,), values=(
                    c.usuario.nombre if c.usuario else "",
                    c.abierto_en.strftime("%d/%m/%Y %H:%M") if c.abierto_en else "",
                    c.cerrado_en.strftime("%d/%m/%Y %H:%M") if c.cerrado_en else "—",
                    dur_str,
                    num_v,
                    f"${c_tv:.2f}",
                    f"${c_ef:.2f}",
                    f"${c_tj:.2f}",
                    f"${c_tr:.2f}",
                    f"${c.monto_apertura:.2f}",
                    f"${c.monto_cierre:.2f}" if c.cerrado_en else "—",
                    dif_str,
                    estado,
                ))
        finally:
            db.close()

    def _exportar_excel(self):
        desde, hasta = self._get_fechas()
        if not desde:
            return

        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            initialfile=f"reporte_{date.today().strftime('%Y%m%d')}.xlsx"
        )
        if not path:
            return

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Ventas"

            ws.merge_cells("A1:I1")
            ws["A1"] = f"Reporte de Ventas — {desde.strftime('%d/%m/%Y')} al {hasta.strftime('%d/%m/%Y')}"
            ws["A1"].font = Font(bold=True, size=14)
            ws["A1"].alignment = Alignment(horizontal="center")

            headers = ["Folio", "Fecha", "Cajero", "Cliente", "Subtotal", "IVA", "Total", "Pago", "Estado"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")

            db = get_db_session()
            try:
                ventas = db.query(Venta).filter(
                    Venta.creado_en >= desde, Venta.creado_en <= hasta,
                    Venta.eliminado.is_not(True),
                ).order_by(Venta.creado_en.desc()).all()

                for row, v in enumerate(ventas, 4):
                    ws.cell(row=row, column=1, value=v.folio or v.id)
                    ws.cell(row=row, column=2, value=v.creado_en.strftime("%d/%m/%Y %H:%M") if v.creado_en else "")
                    ws.cell(row=row, column=3, value=v.usuario.nombre if v.usuario else "")
                    ws.cell(row=row, column=4, value=v.cliente.nombre if v.cliente else "Público")
                    ws.cell(row=row, column=5, value=v.subtotal)
                    ws.cell(row=row, column=6, value=v.iva)
                    ws.cell(row=row, column=7, value=v.total)
                    ws.cell(row=row, column=8, value=v.metodo_pago.value)
                    ws.cell(row=row, column=9, value=v.estado.value)
            finally:
                db.close()

            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            wb.save(path)
            messagebox.showinfo("Exportado", f"Reporte guardado en:\n{path}")
        except ImportError:
            messagebox.showerror("Error", "Instala openpyxl: pip install openpyxl")
        except Exception as e:
            messagebox.showerror("Error", str(e))

    # ── Vencimientos tab ──────────────────────────────────────────────────────

    def _build_vencimientos_tab(self):
        tab = self.tabs.tab("Vencimientos")
        tab.grid_columnconfigure(0, weight=1)
        tab.grid_rowconfigure(1, weight=1)

        # Filter bar
        bar = ctk.CTkFrame(tab, corner_radius=10, fg_color=CARD_BG,
                           border_width=1, border_color=BORDER)
        bar.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        bar.grid_columnconfigure(1, weight=1)

        self.var_venc_filtro = tk.StringVar(value="Todos")
        ctk.CTkSegmentedButton(
            bar,
            values=["Todos", "Vencidos", "Por Vencer", "Vigentes"],
            variable=self.var_venc_filtro,
            command=lambda v: self._load_vencimientos(),
        ).grid(row=0, column=0, padx=14, pady=10, sticky="w")

        ctk.CTkButton(
            bar, text="↺ Actualizar", width=100, height=32, corner_radius=8,
            fg_color=BLUE_L, hover_color="#DBEAFE", text_color=BLUE,
            font=ctk.CTkFont(size=11, weight="bold"),
            command=self._load_vencimientos,
        ).grid(row=0, column=2, padx=14, pady=10, sticky="e")

        # Treeview
        sty = self._make_tree_style("Venc")
        style = ttk.Style()
        style.configure("Venc.Treeview", rowheight=28)

        frame = ctk.CTkFrame(tab, corner_radius=10, fg_color=CARD_BG,
                             border_width=1, border_color=BORDER)
        frame.grid(row=1, column=0, sticky="nsew")
        frame.grid_columnconfigure(0, weight=1)
        frame.grid_rowconfigure(0, weight=1)

        cols = ("producto", "lote", "vencimiento", "cantidad", "dias", "estado")
        self.venc_tree = ttk.Treeview(frame, columns=cols, show="headings", style=sty)
        for col, heading, width, anchor in [
            ("producto",    "Producto",         260, "w"),
            ("lote",        "Nº Lote",          130, "center"),
            ("vencimiento", "Fecha Vencim.",    130, "center"),
            ("cantidad",    "Cantidad",          80, "center"),
            ("dias",        "Días Rest.",        90, "center"),
            ("estado",      "Estado",           120, "center"),
        ]:
            self.venc_tree.heading(col, text=heading)
            self.venc_tree.column(col, width=width, anchor=anchor)

        self.venc_tree.tag_configure("vencido",    foreground="#DC2626", background="#FEF2F2")
        self.venc_tree.tag_configure("por_vencer", foreground="#D97706", background="#FFFBEB")
        self.venc_tree.tag_configure("ok",         foreground="#16A34A")

        scroll_y = ttk.Scrollbar(frame, orient="vertical", command=self.venc_tree.yview)
        scroll_x = ttk.Scrollbar(frame, orient="horizontal", command=self.venc_tree.xview)
        self.venc_tree.configure(yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set)
        self.venc_tree.grid(row=0, column=0, sticky="nsew")
        scroll_y.grid(row=0, column=1, sticky="ns")
        scroll_x.grid(row=1, column=0, sticky="ew")

        self.lbl_venc_status = ctk.CTkLabel(tab, text="", font=ctk.CTkFont(size=11),
                                             text_color=MUTED)
        self.lbl_venc_status.grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))

        self._load_vencimientos()

    def _load_vencimientos(self):
        for row in self.venc_tree.get_children():
            self.venc_tree.delete(row)

        hoy = date.today()
        alerta = hoy + timedelta(days=cfg.EXPIRY_ALERT_DAYS)
        filtro = self.var_venc_filtro.get()

        db = get_db_session()
        try:
            lotes = (db.query(Lote)
                     .join(Producto, Lote.producto_id == Producto.id)
                     .filter(Producto.activo == True, Lote.cantidad > 0)
                     .order_by(Lote.fecha_vencimiento)
                     .all())

            count = 0
            for lote in lotes:
                fv = lote.fecha_vencimiento
                if fv:
                    dias = (fv - hoy).days
                    fv_str = fv.strftime("%d/%m/%Y")
                    dias_str = str(dias) if dias >= 0 else f"Hace {-dias}"
                    if fv <= hoy:
                        tag = "vencido";    estado = "⛔ Vencido"
                        if filtro in ("Por Vencer", "Vigentes"):
                            continue
                    elif fv <= alerta:
                        tag = "por_vencer"; estado = "⚠ Por vencer"
                        if filtro in ("Vencidos", "Vigentes"):
                            continue
                    else:
                        tag = "ok";         estado = "✓ Vigente"
                        if filtro in ("Vencidos", "Por Vencer"):
                            continue
                else:
                    fv_str = "Sin fecha"; dias_str = "-"; tag = "ok"; estado = "Sin fecha"
                    if filtro in ("Vencidos", "Por Vencer"):
                        continue

                nombre = lote.producto.nombre if lote.producto else "?"
                self.venc_tree.insert("", "end", values=(
                    nombre,
                    lote.numero_lote or "S/N",
                    fv_str,
                    lote.cantidad,
                    dias_str,
                    estado,
                ), tags=(tag,))
                count += 1

            self.lbl_venc_status.configure(text=f"Mostrando {count} lotes")
        finally:
            db.close()

    # ── Admin purge ───────────────────────────────────────────────────────────

    def _dlg_purgar(self):
        dlg = ctk.CTkToplevel(self)
        dlg.title("⚠️ Purgar datos")
        dlg.geometry("440x340")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width()  - 440) // 2
        y = self.winfo_rooty() + (self.winfo_height() - 340) // 2
        dlg.geometry(f"440x340+{x}+{y}")

        ctk.CTkLabel(dlg, text="⚠️  Eliminar datos",
                     font=ctk.CTkFont(size=15, weight="bold"),
                     text_color="#EF4444").pack(pady=(20, 4))
        ctk.CTkLabel(dlg, text="Selecciona qué deseas eliminar (local + Turso).",
                     font=ctk.CTkFont(size=11), text_color="#94A3B8").pack(pady=(0, 16))

        # Option 1
        ctk.CTkButton(
            dlg,
            text="🗑  Eliminar ventas, historial y cierres",
            height=40, corner_radius=10,
            fg_color="#F97316", hover_color="#C2410C", text_color="white",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=lambda: self._pedir_pin(
                dlg,
                "Eliminar ventas, historial y cierres",
                "Se borrarán ventas, movimientos, auditoría y cortes.\n"
                "Conserva productos, clientes y proveedores.",
                self._purgar_ventas,
            ),
        ).pack(fill="x", padx=24, pady=(0, 8))
        ctk.CTkLabel(
            dlg,
            text="Conserva: productos, clientes, proveedores, categorías.",
            font=ctk.CTkFont(size=10), text_color="#94A3B8",
        ).pack(anchor="w", padx=28, pady=(0, 12))

        ctk.CTkFrame(dlg, height=1, fg_color="#EF4444").pack(fill="x", padx=24, pady=(0, 12))

        # Option 2
        ctk.CTkButton(
            dlg,
            text="💀  Eliminar TODO sin dejar nada",
            height=40, corner_radius=10,
            fg_color="#EF4444", hover_color="#7F1D1D", text_color="white",
            font=ctk.CTkFont(size=12, weight="bold"),
            command=lambda: self._pedir_pin(
                dlg,
                "Eliminar TODOS los registros",
                "Se eliminará absolutamente todo excepto usuarios.\nACCIÓN IRREVERSIBLE.",
                self._purgar_todo,
            ),
        ).pack(fill="x", padx=24, pady=(0, 4))
        ctk.CTkLabel(
            dlg,
            text="Elimina: productos, ventas, clientes, compras, todo.",
            font=ctk.CTkFont(size=10), text_color="#94A3B8",
        ).pack(anchor="w", padx=28, pady=(0, 12))

        ctk.CTkButton(
            dlg, text="Cancelar", height=32, corner_radius=8,
            fg_color="transparent", border_width=1, border_color="#E2E8F0",
            text_color="#94A3B8", command=dlg.destroy,
        ).pack(pady=(4, 0))

    def _pedir_pin(self, parent_dlg, titulo, descripcion, on_confirm):
        parent_dlg.destroy()
        pin_dlg = ctk.CTkToplevel(self)
        pin_dlg.title(f"⚠️ {titulo}")
        pin_dlg.geometry("420x260")
        pin_dlg.resizable(False, False)
        pin_dlg.grab_set()
        pin_dlg.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width()  - 420) // 2
        y = self.winfo_rooty() + (self.winfo_height() - 260) // 2
        pin_dlg.geometry(f"420x260+{x}+{y}")

        ctk.CTkLabel(pin_dlg, text=f"⚠️  {titulo}",
                     font=ctk.CTkFont(size=13, weight="bold"),
                     text_color="#EF4444", wraplength=380).pack(pady=(18, 4))
        ctk.CTkLabel(pin_dlg, text=descripcion,
                     font=ctk.CTkFont(size=11), text_color="#94A3B8",
                     wraplength=380, justify="center").pack(pady=(0, 10))
        ctk.CTkLabel(pin_dlg, text="PIN de administrador:",
                     font=ctk.CTkFont(size=12)).pack(pady=(0, 4))

        entry = ctk.CTkEntry(pin_dlg, width=150, height=36, justify="center", show="●")
        entry.pack(pady=(0, 4))
        entry.focus()

        lbl_err = ctk.CTkLabel(pin_dlg, text="", text_color="#EF4444",
                               font=ctk.CTkFont(size=11))
        lbl_err.pack(pady=(0, 6))

        def _ok(event=None):
            if entry.get().strip() != _get_admin_pin():
                lbl_err.configure(text="PIN incorrecto")
                entry.delete(0, "end")
                return
            pin_dlg.destroy()
            on_confirm()

        entry.bind("<Return>", _ok)
        ctk.CTkButton(pin_dlg, text="Confirmar", height=34,
                      fg_color="#EF4444", hover_color="#B91C1C", text_color="white",
                      command=_ok).pack(pady=(0, 6))
        ctk.CTkButton(pin_dlg, text="Cancelar", height=30,
                      fg_color="transparent", border_width=1, border_color="#E2E8F0",
                      text_color="#94A3B8", command=pin_dlg.destroy).pack()

    def _purgar_ventas(self):
        import threading
        from app.database.sync_service import purgar_ventas_historial_cierres
        from app.ui import toast

        def _run():
            try:
                purgar_ventas_historial_cierres()
                self.after(0, lambda: (
                    toast.show("Ventas, historial y cierres eliminados", kind="success", duration=5000),
                    self._generar(),
                ))
            except Exception as exc:
                self.after(0, lambda e=exc: toast.show(f"Error: {e}", kind="error", duration=7000))

        toast.show("Eliminando ventas, historial y cierres…", kind="warning", duration=15000)
        threading.Thread(target=_run, daemon=True, name="PurgarVentasR").start()

    def _purgar_todo(self):
        import threading
        from app.database.sync_service import purgar_todos_los_datos
        from app.ui import toast

        def _run():
            try:
                purgar_todos_los_datos()
                self.after(0, lambda: (
                    toast.show("Todos los registros eliminados", kind="success", duration=5000),
                    self._generar(),
                ))
            except Exception as exc:
                self.after(0, lambda e=exc: toast.show(f"Error: {e}", kind="error", duration=7000))

        toast.show("Eliminando todos los registros…", kind="warning", duration=15000)
        threading.Thread(target=_run, daemon=True, name="PurgarTodoR").start()

    # ── Individual sale delete (admin only) ──────────────────────────────────

    def _ctx_menu_venta(self, event):
        iid = self.ventas_tree.identify_row(event.y)
        if not iid:
            return
        self.ventas_tree.selection_set(iid)
        menu = tk.Menu(self, tearoff=0)
        menu.add_command(label="👁  Ver detalle", command=lambda: self._ver_detalle_venta(None))
        menu.add_command(label="🗑  Eliminar venta", command=self._eliminar_venta_seleccionada)
        try:
            menu.post(event.x_root, event.y_root)
        finally:
            menu.grab_release()

    def _eliminar_venta_seleccionada(self):
        sel = self.ventas_tree.selection()
        if not sel:
            messagebox.showwarning("Sin selección", "Selecciona una venta de la lista primero.")
            return
        venta_id = self._venta_ids.get(sel[0])
        if not venta_id:
            return
        vals = self.ventas_tree.item(sel[0], "values")
        folio = vals[0] if vals else str(venta_id)
        self._pedir_pin_eliminar_venta(venta_id, folio)

    def _pedir_pin_eliminar_venta(self, venta_id: int, folio: str):
        dlg = ctk.CTkToplevel(self)
        dlg.title("Eliminar venta")
        dlg.geometry("400x250")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width()  - 400) // 2
        y = self.winfo_rooty() + (self.winfo_height() - 250) // 2
        dlg.geometry(f"400x250+{x}+{y}")

        ctk.CTkLabel(dlg, text=f"Eliminar venta {folio}",
                     font=ctk.CTkFont(size=14, weight="bold"),
                     text_color="#EF4444").pack(pady=(20, 4))
        ctk.CTkLabel(dlg,
                     text="El inventario se restaurará automáticamente.\nIngresa el PIN de administrador:",
                     font=ctk.CTkFont(size=11), text_color=MUTED,
                     justify="center").pack(pady=(0, 8))

        entry = ctk.CTkEntry(dlg, width=140, height=36, justify="center", show="●")
        entry.pack(pady=(0, 4))
        entry.focus()

        lbl_err = ctk.CTkLabel(dlg, text="", text_color="#EF4444",
                               font=ctk.CTkFont(size=11))
        lbl_err.pack(pady=(0, 6))

        def _ok(event=None):
            if entry.get().strip() != _get_admin_pin():
                lbl_err.configure(text="PIN incorrecto")
                entry.delete(0, "end")
                return
            dlg.destroy()
            self._do_eliminar_venta(venta_id)

        entry.bind("<Return>", _ok)
        ctk.CTkButton(dlg, text="Confirmar", height=34,
                      fg_color="#EF4444", hover_color="#B91C1C", text_color="white",
                      command=_ok).pack(pady=(0, 6))
        ctk.CTkButton(dlg, text="Cancelar", height=30,
                      fg_color="transparent", border_width=1, border_color=BORDER,
                      text_color=MUTED, command=dlg.destroy).pack()

    def _do_eliminar_venta(self, venta_id: int):
        import threading
        from app.database.sync_service import eliminar_venta
        from app.ui import toast

        def _run():
            try:
                result = eliminar_venta(venta_id)
                folio = result.get("folio", str(venta_id))
                self.after(0, lambda: (
                    toast.show(f"Venta {folio} eliminada — stock restaurado", kind="success", duration=4000),
                    self._generar(),
                ))
            except Exception as exc:
                self.after(0, lambda e=exc: toast.show(f"Error al eliminar: {e}", kind="error", duration=7000))

        threading.Thread(target=_run, daemon=True, name="EliminarVenta").start()

    def _ver_detalle_venta(self, event=None):
        sel = self.ventas_tree.selection()
        if not sel:
            return
        venta_id = self._venta_ids.get(sel[0])
        if not venta_id:
            return

        # Eagerly load all data before closing DB
        db = get_db_session()
        try:
            v = db.query(Venta).filter(
                Venta.id == venta_id, Venta.eliminado.is_not(True)
            ).first()
            if not v:
                messagebox.showwarning("Error", "Venta no encontrada.")
                return
            folio       = v.folio or str(venta_id)
            fecha_str   = v.creado_en.strftime("%d/%m/%Y %H:%M") if v.creado_en else ""
            cajero      = v.usuario.nombre if v.usuario else "—"
            cliente     = v.cliente.nombre if v.cliente else "Público General"
            metodo      = v.metodo_pago.value.capitalize()
            estado_str  = v.estado.value.capitalize()
            subtotal    = v.subtotal
            iva         = v.iva
            total       = v.total
            items_data  = [
                (
                    item.producto.nombre if item.producto else f"(Producto #{item.producto_id})",
                    item.cantidad,
                    item.precio_unitario,
                    item.subtotal,
                )
                for item in v.items
            ]
        finally:
            db.close()

        win = ctk.CTkToplevel(self)
        win.title(f"Detalle Venta — {folio}")
        win.geometry("580x540")
        win.resizable(False, True)
        win.grab_set()
        win.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width()  - 580) // 2
        y = self.winfo_rooty() + (self.winfo_height() - 540) // 2
        win.geometry(f"580x540+{x}+{y}")

        # ── Header ────────────────────────────────────────────────────────────
        ctk.CTkLabel(win, text=f"🧾  Venta {folio}",
                     font=ctk.CTkFont(size=16, weight="bold"),
                     text_color=TEXT).pack(pady=(18, 2))
        ctk.CTkLabel(win, text=fecha_str,
                     font=ctk.CTkFont(size=11), text_color=MUTED).pack()

        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=10)

        # ── Info rows ─────────────────────────────────────────────────────────
        def info_row(label, value, color=TEXT):
            f = ctk.CTkFrame(win, fg_color="transparent")
            f.pack(fill="x", padx=28, pady=2)
            ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=11),
                         text_color=MUTED, anchor="w", width=160).pack(side="left")
            ctk.CTkLabel(f, text=value, font=ctk.CTkFont(size=11, weight="bold"),
                         text_color=color, anchor="w").pack(side="left")

        info_row("Cajero:", cajero)
        info_row("Cliente:", cliente)
        info_row("Método de pago:", metodo, BLUE)
        info_row("Estado:", estado_str, GREEN)

        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=10)
        ctk.CTkLabel(win, text="Productos vendidos",
                     font=ctk.CTkFont(size=12, weight="bold"),
                     text_color=TEXT, anchor="w").pack(fill="x", padx=28, pady=(0, 6))

        # ── Items treeview ────────────────────────────────────────────────────
        tree_frame = ctk.CTkFrame(win, fg_color=CARD_BG, corner_radius=8,
                                   border_width=1, border_color=BORDER)
        tree_frame.pack(fill="both", expand=True, padx=20, pady=(0, 8))
        tree_frame.grid_columnconfigure(0, weight=1)
        tree_frame.grid_rowconfigure(0, weight=1)

        sty = self._make_tree_style("Det")
        item_cols = ("producto", "cantidad", "precio", "subtotal")
        items_tree = ttk.Treeview(tree_frame, columns=item_cols,
                                   show="headings", style=sty, height=6)
        for col, heading, width, anchor in [
            ("producto",  "Producto",  280, "w"),
            ("cantidad",  "Cantidad",   70, "center"),
            ("precio",    "Precio",     90, "e"),
            ("subtotal",  "Subtotal",   90, "e"),
        ]:
            items_tree.heading(col, text=heading)
            items_tree.column(col, width=width, anchor=anchor)
        items_tree.tag_configure("even", background="#F8FAFF")
        items_tree.tag_configure("odd",  background="#FFFFFF")

        sc = ttk.Scrollbar(tree_frame, orient="vertical", command=items_tree.yview)
        items_tree.configure(yscrollcommand=sc.set)
        items_tree.grid(row=0, column=0, sticky="nsew", padx=(4, 0), pady=4)
        sc.grid(row=0, column=1, sticky="ns", pady=4)

        for idx, (nombre, qty, precio, sub) in enumerate(items_data):
            tag = "even" if idx % 2 == 0 else "odd"
            items_tree.insert("", "end", tags=(tag,), values=(
                nombre, qty, f"${precio:.2f}", f"${sub:.2f}"
            ))

        # ── Totals ────────────────────────────────────────────────────────────
        ctk.CTkFrame(win, height=1, fg_color=BORDER).pack(fill="x", padx=20, pady=(4, 4))
        tf = ctk.CTkFrame(win, fg_color="transparent")
        tf.pack(fill="x", padx=28)

        def total_row(label, value, color=TEXT, bold=False):
            f = ctk.CTkFrame(tf, fg_color="transparent")
            f.pack(fill="x", pady=1)
            ctk.CTkLabel(f, text=label, font=ctk.CTkFont(size=11),
                         text_color=MUTED, anchor="w").pack(side="left")
            ctk.CTkLabel(f, text=value,
                         font=ctk.CTkFont(size=13 if bold else 11, weight="bold"),
                         text_color=color, anchor="e").pack(side="right")

        total_row("Subtotal:", f"${subtotal:.2f}")
        total_row("IVA:", f"${iva:.2f}")
        total_row("Total:", f"${total:.2f}", BLUE, bold=True)

        # ── Buttons ───────────────────────────────────────────────────────────
        bf = ctk.CTkFrame(win, fg_color="transparent")
        bf.pack(fill="x", padx=20, pady=(8, 16))

        if self.user.rol == RolUsuario.admin:
            def _eliminar():
                win.destroy()
                self._pedir_pin_eliminar_venta(venta_id, folio)
            ctk.CTkButton(
                bf, text="🗑  Eliminar venta", width=160, height=34, corner_radius=8,
                fg_color="#FEF2F2", hover_color="#FEE2E2", text_color="#EF4444",
                border_width=1, border_color="#FECACA",
                font=ctk.CTkFont(size=11, weight="bold"),
                command=_eliminar,
            ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            bf, text="Cerrar", width=90, height=34, corner_radius=8,
            fg_color="transparent", border_width=1, border_color=BORDER,
            text_color=MUTED, command=win.destroy,
        ).pack(side="right")

    def on_show(self):
        self._generar()
