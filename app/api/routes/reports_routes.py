from fastapi import APIRouter, Depends, HTTPException, Query, BackgroundTasks
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import date, datetime, timedelta
from sqlalchemy import func
from sqlalchemy.orm import joinedload, selectinload
import io, csv, os, tempfile

from app.database.connection import get_db_session
from app.database.models import Venta, ItemVenta, Producto, EstadoVenta, CortesCaja, Lote, MovimientoStock, TipoMovimiento
from app.api.routes.auth_routes import get_current_api_user

router = APIRouter()


def _require_admin(payload: dict):
    if payload.get("rol") != "admin":
        raise HTTPException(status_code=403, detail="Solo administradores")


def _rango(fecha_inicio: date, fecha_fin: date):
    return (
        datetime.combine(fecha_inicio, datetime.min.time()),
        datetime.combine(fecha_fin,    datetime.max.time()),
    )


@router.get("/resumen")
def resumen(
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    db = get_db_session()
    try:
        fi, ff = _rango(fecha_inicio, fecha_fin)
        ventas = (
            db.query(Venta)
            .filter(Venta.creado_en >= fi, Venta.creado_en <= ff,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True))
            .all()
        )
        total = sum(v.total for v in ventas)
        num   = len(ventas)

        por_dia: dict[str, float] = {}
        for v in ventas:
            d = v.creado_en.date().isoformat() if v.creado_en else "?"
            por_dia[d] = por_dia.get(d, 0.0) + v.total

        mejor_dia   = max(por_dia, key=por_dia.get) if por_dia else None
        mejor_monto = por_dia[mejor_dia] if mejor_dia else 0.0

        # Cost of goods sold for period
        venta_ids = [v.id for v in ventas]
        if venta_ids:
            cost_rows = (
                db.query(ItemVenta.cantidad, Producto.precio_compra)
                .join(Producto, ItemVenta.producto_id == Producto.id)
                .filter(ItemVenta.venta_id.in_(venta_ids))
                .all()
            )
            total_costo = sum(r.cantidad * (r.precio_compra or 0.0) for r in cost_rows)
        else:
            total_costo = 0.0
        # Partial devoluciones in period (full returns excluded via estado=devolucion filter)
        dev_movs = db.query(MovimientoStock).filter(
            MovimientoStock.tipo == TipoMovimiento.devolucion,
            MovimientoStock.referencia_tipo == "devolucion",
            MovimientoStock.creado_en >= fi,
            MovimientoStock.creado_en <= ff,
        ).all()
        total_devoluciones = 0.0
        if dev_movs:
            venta_ids_dev = {mov.referencia_id for mov in dev_movs}
            # Una sola consulta por lote en vez de una por cada movimiento (evita N+1).
            precios_por_par = {
                (i.venta_id, i.producto_id): i.precio_unitario
                for i in db.query(ItemVenta).filter(ItemVenta.venta_id.in_(venta_ids_dev)).all()
            }
            for mov in dev_movs:
                precio = precios_por_par.get((mov.referencia_id, mov.producto_id))
                if precio is not None:
                    total_devoluciones += precio * mov.cantidad
        ventas_netas = total - total_devoluciones
        ganancia = ventas_netas - total_costo

        return {
            "total":               total,
            "total_devoluciones":  round(total_devoluciones, 2),
            "ventas_netas":        round(ventas_netas, 2),
            "num_ventas":          num,
            "ticket_promedio":     total / num if num else 0.0,
            "mejor_dia_fecha":     mejor_dia,
            "mejor_dia_monto":     mejor_monto,
            "efectivo":            sum(v.total for v in ventas if v.metodo_pago.value == "efectivo"),
            "tarjeta":             sum(v.total for v in ventas if v.metodo_pago.value == "tarjeta"),
            "transferencia":       sum(v.total for v in ventas if v.metodo_pago.value == "transferencia"),
            "total_costo":         total_costo,
            "ganancia":            round(ganancia, 2),
            "por_dia": [{"fecha": k, "total": v} for k, v in sorted(por_dia.items())],
        }
    finally:
        db.close()


@router.get("/top-productos")
def top_productos(
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    limite: int = Query(10, ge=1, le=50),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    db = get_db_session()
    try:
        fi, ff = _rango(fecha_inicio, fecha_fin)
        rows = (
            db.query(
                Producto.nombre,
                func.sum(ItemVenta.cantidad).label("total_vendido"),
                func.sum(ItemVenta.subtotal).label("total_ingreso"),
            )
            .join(ItemVenta, ItemVenta.producto_id == Producto.id)
            .join(Venta,     Venta.id == ItemVenta.venta_id)
            .filter(
                Venta.creado_en >= fi,
                Venta.creado_en <= ff,
                Venta.estado == EstadoVenta.completada,
                Venta.eliminado.is_not(True),
            )
            .group_by(Producto.id, Producto.nombre)
            .order_by(func.sum(ItemVenta.cantidad).desc())
            .limit(limite)
            .all()
        )
        return [
            {"nombre": r.nombre, "cantidad": r.total_vendido or 0, "ingreso": r.total_ingreso or 0.0}
            for r in rows
        ]
    finally:
        db.close()


@router.get("/inventario")
def reporte_inventario(payload: dict = Depends(get_current_api_user)):
    _require_admin(payload)
    db = get_db_session()
    try:
        productos = (
            db.query(Producto)
            .filter(Producto.activo == True)
            .order_by(Producto.stock.asc(), Producto.nombre.asc())
            .all()
        )
        return [
            {
                "id":           p.id,
                "nombre":       p.nombre,
                "stock":        p.stock,
                "stock_minimo": p.stock_minimo,
                "precio_venta": p.precio_venta,
                "bajo_stock":   p.stock <= p.stock_minimo,
            }
            for p in productos
        ]
    finally:
        db.close()


@router.get("/cortes")
def cortes_caja(
    fecha_inicio:   date = Query(...),
    fecha_fin:      date = Query(...),
    ocultar_vacios: bool = Query(True),   # hide 0-sale, $0 phantom shifts by default
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    db = get_db_session()
    try:
        fi, ff = _rango(fecha_inicio, fecha_fin)
        q = (
            db.query(CortesCaja)
            .filter(CortesCaja.abierto_en >= fi, CortesCaja.abierto_en <= ff)
        )
        if ocultar_vacios:
            # Exclude shifts that have 0 ventas AND $0 total (phantom / test shifts)
            from sqlalchemy import or_
            q = q.filter(
                or_(
                    CortesCaja.num_ventas > 0,
                    CortesCaja.total_ventas > 0,
                    CortesCaja.cerrado_en == None,  # keep open shifts even if 0 so far
                )
            )
        cortes = q.order_by(CortesCaja.abierto_en.desc()).all()
        result = []
        for c in cortes:
            dur_min = None
            if c.cerrado_en and c.abierto_en:
                dur_min = int((c.cerrado_en - c.abierto_en).total_seconds() / 60)
            ef  = c.total_efectivo or 0.0
            ape = c.monto_apertura or 0.0
            esperado = ape + ef
            dif = (c.monto_cierre - esperado) if c.monto_cierre is not None else None
            result.append({
                "id":               c.id,
                "cajero":           c.usuario.nombre if c.usuario else "",
                "abierto_en":       c.abierto_en.isoformat() if c.abierto_en else None,
                "cerrado_en":       c.cerrado_en.isoformat() if c.cerrado_en else None,
                "duracion_min":     dur_min,
                "num_ventas":       c.num_ventas or 0,
                "total_ventas":     c.total_ventas or 0.0,
                "total_efectivo":   ef,
                "total_tarjeta":    c.total_tarjeta or 0.0,
                "total_transferencia": c.total_transferencia or 0.0,
                "monto_apertura":   ape,
                "monto_cierre":     c.monto_cierre,
                "esperado_caja":    esperado,
                "diferencia":       dif,
                "notas":            c.notas or "",
                "abierto":          c.cerrado_en is None,
            })
        return result
    finally:
        db.close()


@router.get("/vencimientos")
def vencimientos(
    filtro: str = Query("todos"),   # todos | vencidos | por_vencer | vigentes
    dias:   int = Query(30),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    db = get_db_session()
    try:
        hoy = date.today()
        alerta = hoy + timedelta(days=dias)
        lotes = (
            db.query(Lote)
            .join(Producto, Lote.producto_id == Producto.id)
            .filter(Producto.activo == True, Lote.cantidad > 0)
            .order_by(Lote.fecha_vencimiento)
            .all()
        )
        result = []
        for l in lotes:
            fv = l.fecha_vencimiento
            if fv:
                dias_rest = (fv - hoy).days
                if fv < hoy:
                    estado = "vencido"
                elif fv <= alerta:
                    estado = "por_vencer"
                else:
                    estado = "vigente"
            else:
                dias_rest = None
                estado = "sin_fecha"

            if filtro == "vencidos"    and estado != "vencido":    continue
            if filtro == "por_vencer"  and estado != "por_vencer": continue
            if filtro == "vigentes"    and estado not in ("vigente","sin_fecha"): continue

            result.append({
                "lote_id":           l.id,
                "producto_id":       l.producto_id,
                "producto":          l.producto.nombre if l.producto else "",
                "numero_lote":       l.numero_lote or "",
                "fecha_vencimiento": fv.isoformat() if fv else None,
                "cantidad":          l.cantidad,
                "dias_restantes":    dias_rest,
                "estado":            estado,
            })
        return result
    finally:
        db.close()


@router.get("/export-csv")
def export_csv(
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    db = get_db_session()
    try:
        fi, ff = _rango(fecha_inicio, fecha_fin)
        from app.database.models import EstadoVenta as _EV
        ventas = (
            db.query(Venta)
            .options(joinedload(Venta.items))
            .filter(
                Venta.creado_en >= fi,
                Venta.creado_en <= ff,
                Venta.estado == _EV.completada,
                Venta.eliminado.is_not(True),
            )
            .order_by(Venta.creado_en)
            .all()
        )

        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(["Folio", "Fecha", "Subtotal", "Descuento", "IVA", "Total",
                    "Metodo Pago", "Monto Pagado", "Cambio", "Estado", "N Articulos"])
        for v in ventas:
            w.writerow([
                v.folio or v.id,
                v.creado_en.strftime("%Y-%m-%d %H:%M") if v.creado_en else "",
                round(v.subtotal, 2),
                round(v.descuento, 2),
                round(v.iva, 2),
                round(v.total, 2),
                v.metodo_pago.value,
                round(v.monto_pagado, 2),
                round(v.cambio, 2),
                v.estado.value,
                len(v.items),
            ])

        buf.seek(0)
        filename = f"ventas_{fecha_inicio}_{fecha_fin}.csv"
        return StreamingResponse(
            iter([buf.getvalue().encode("utf-8-sig")]),  # utf-8-sig = Excel-compatible BOM
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )
    finally:
        db.close()


MESES_ES = ["Enero","Febrero","Marzo","Abril","Mayo","Junio",
            "Julio","Agosto","Septiembre","Octubre","Noviembre","Diciembre"]


def _build_cierre_mensual_pdf_bytes(mes: int, anio: int) -> bytes:  # noqa: C901
    """Generate RESICO monthly closing PDF — constancia/INFONAVIT style."""
    import calendar
    import app.config as cfg
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT, TA_JUSTIFY
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, KeepTogether,
                                    HRFlowable, Flowable, PageBreak)

    # ── Signature-line Flowable (canvas) ─────────────────────────────────────
    class _SigLine(Flowable):
        def __init__(self, w, label, subtext=""):
            super().__init__()
            self.width = w; self.height = 1.6 * cm
            self.label = label; self.subtext = subtext
        def draw(self):
            c = self.canv; cx = self.width / 2
            c.saveState()
            c.setStrokeColor(colors.HexColor("#1d2140"))
            c.setLineWidth(0.6)
            pad = 0.8 * cm
            c.line(pad, 0.85 * cm, self.width - pad, 0.85 * cm)
            c.setFont("Helvetica-Bold", 6.5)
            c.setFillColor(colors.HexColor("#1d2140"))
            c.drawCentredString(cx, 0.55 * cm, self.label)
            if self.subtext:
                c.setFont("Helvetica", 6)
                c.setFillColor(colors.HexColor("#64748B"))
                c.drawCentredString(cx, 0.32 * cm, self.subtext)
            c.restoreState()
        def wrap(self, aw, ah): return self.width, self.height

    # ── Palette (constancia: almost monochrome, navy accent) ─────────────────
    C_NAVY  = colors.HexColor("#0D1B3E")
    C_NAVY2 = colors.HexColor("#1d2140")
    C_DARK  = colors.HexColor("#0F172A")
    C_MUTED = colors.HexColor("#64748B")
    C_LINE  = colors.HexColor("#CBD5E1")
    C_LIGHT = colors.HexColor("#F1F5F9")
    C_WHITE = colors.white
    C_BLACK = colors.black
    C_GOLD  = colors.HexColor("#92400E")  # dark amber for totals

    W, H = letter
    _, last_day = calendar.monthrange(anio, mes)
    fecha_inicio = date(anio, mes, 1)
    fecha_fin    = date(anio, mes, last_day)
    fi, ff       = _rango(fecha_inicio, fecha_fin)
    nombre_mes   = MESES_ES[mes - 1]
    per_label    = f"{nombre_mes} {anio}".upper()
    now_str      = datetime.now().strftime("%d/%m/%Y %H:%M")
    rfc_val      = cfg.PHARMACY_RFC

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    # Constancia typography
    s_title  = sty("_t",  fontSize=13, textColor=C_NAVY,  fontName="Helvetica-Bold", leading=16, alignment=TA_CENTER)
    s_sub    = sty("_s",  fontSize=8,  textColor=C_MUTED, fontName="Helvetica",      leading=12, alignment=TA_CENTER)
    s_fld_l  = sty("_fl", fontSize=6,  textColor=C_MUTED, fontName="Helvetica-Bold", leading=8)
    s_fld_v  = sty("_fv", fontSize=9,  textColor=C_DARK,  fontName="Helvetica-Bold", leading=13)
    s_fld_vr = sty("_fr", fontSize=9,  textColor=C_DARK,  fontName="Helvetica-Bold", leading=13, alignment=TA_RIGHT)
    s_body   = sty("_b",  fontSize=8,  textColor=C_DARK,  fontName="Helvetica",      leading=12)
    s_th     = sty("_h",  fontSize=7.5,textColor=C_WHITE, fontName="Helvetica-Bold", leading=11, alignment=TA_CENTER)
    s_th_l   = sty("_hl", fontSize=7.5,textColor=C_WHITE, fontName="Helvetica-Bold", leading=11)
    s_td     = sty("_d",  fontSize=7.5,textColor=C_DARK,  fontName="Helvetica",      leading=11)
    s_td_r   = sty("_dr", fontSize=7.5,textColor=C_DARK,  fontName="Helvetica",      leading=11, alignment=TA_RIGHT)
    s_td_c   = sty("_dc", fontSize=7.5,textColor=C_DARK,  fontName="Helvetica",      leading=11, alignment=TA_CENTER)
    s_td_b   = sty("_db", fontSize=7.5,textColor=C_DARK,  fontName="Helvetica-Bold", leading=11)
    s_td_br  = sty("_dbr",fontSize=7.5,textColor=C_DARK,  fontName="Helvetica-Bold", leading=11, alignment=TA_RIGHT)
    s_mut    = sty("_m",  fontSize=6.5,textColor=C_MUTED, fontName="Helvetica",      leading=10)
    s_mut_c  = sty("_mc", fontSize=6.5,textColor=C_MUTED, fontName="Helvetica",      leading=10, alignment=TA_CENTER)
    s_mut_r  = sty("_mr", fontSize=6.5,textColor=C_MUTED, fontName="Helvetica",      leading=10, alignment=TA_RIGHT)
    s_legal  = sty("_lg", fontSize=6.5,textColor=C_MUTED, fontName="Helvetica",      leading=9.5, alignment=TA_JUSTIFY)

    cw = W - 3.6 * cm   # narrower margins = more table space, constancia feel

    # ── Header / footer drawn on canvas ──────────────────────────────────────
    def _on_page(canvas, doc):
        canvas.saveState()
        # Navy full-width top bar (constancia style)
        canvas.setFillColor(C_NAVY)
        canvas.rect(0, H - 1.35 * cm, W, 1.35 * cm, fill=1, stroke=0)
        # Thin gold rule below bar
        canvas.setStrokeColor(colors.HexColor("#B45309"))
        canvas.setLineWidth(1.5)
        canvas.line(0, H - 1.35 * cm, W, H - 1.35 * cm)
        # Header text
        canvas.setFont("Helvetica-Bold", 8)
        canvas.setFillColor(C_WHITE)
        canvas.drawString(1.8 * cm, H - 0.85 * cm, cfg.PHARMACY_NAME)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.HexColor("#93C5FD"))
        canvas.drawRightString(W - 1.8 * cm, H - 0.85 * cm,
                               f"RFC: {rfc_val}  ·  RESICO  ·  {per_label}")
        # Footer rule
        canvas.setStrokeColor(C_NAVY)
        canvas.setLineWidth(1.0)
        canvas.line(1.8 * cm, 1.9 * cm, W - 1.8 * cm, 1.9 * cm)
        canvas.setLineWidth(0.3)
        canvas.setStrokeColor(C_LINE)
        canvas.line(1.8 * cm, 1.75 * cm, W - 1.8 * cm, 1.75 * cm)
        canvas.setFont("Helvetica", 6.5)
        canvas.setFillColor(C_MUTED)
        canvas.drawString(1.8 * cm, 1.4 * cm, f"Generado: {now_str}")
        canvas.drawCentredString(W / 2, 1.4 * cm,
                                 f"{cfg.PHARMACY_ADDRESS}  ·  {cfg.PHARMACY_PHONE}")
        canvas.drawRightString(W - 1.8 * cm, 1.4 * cm,
                               f"Página {doc.page}")
        canvas.restoreState()

    # ── Query ─────────────────────────────────────────────────────────────────
    db = get_db_session()
    try:
        ventas = (
            db.query(Venta)
            .options(selectinload(Venta.usuario), selectinload(Venta.cliente))
            .filter(Venta.creado_en >= fi, Venta.creado_en <= ff,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True))
            .order_by(Venta.creado_en)
            .all()
        )
        total_ingresos = sum(v.total for v in ventas)
        num_ventas     = len(ventas)
        ef  = sum(v.total for v in ventas if v.metodo_pago.value == "efectivo")
        tj  = sum(v.total for v in ventas if v.metodo_pago.value == "tarjeta")
        tr  = sum(v.total for v in ventas if v.metodo_pago.value == "transferencia")
        tick_prom = total_ingresos / num_ventas if num_ventas else 0.0
        por_dia: dict[str, dict] = {}
        for v in ventas:
            d = v.creado_en.date().isoformat() if v.creado_en else "?"
            if d not in por_dia:
                por_dia[d] = {"ef": 0.0, "tj": 0.0, "tr": 0.0, "n": 0}
            pv = v.metodo_pago.value
            por_dia[d]["ef"] += v.total if pv == "efectivo"      else 0
            por_dia[d]["tj"] += v.total if pv == "tarjeta"       else 0
            por_dia[d]["tr"] += v.total if pv == "transferencia" else 0
            por_dia[d]["n"]  += 1
        dias_con_vta = len(por_dia)

        def money(v): return f"${v:,.2f}"

        story = []

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE 1 — PORTADA estilo CONSTANCIA (centrada, formal)
        # ═══════════════════════════════════════════════════════════════════
        story.append(Spacer(1, 0.4 * cm))

        # Título centrado grande
        story.append(Paragraph(
            "DECLARACIÓN MENSUAL DE INGRESOS",
            sty("pt", fontSize=15, textColor=C_NAVY, fontName="Helvetica-Bold",
                leading=19, alignment=TA_CENTER),
        ))
        story.append(Paragraph(
            "Régimen Simplificado de Confianza — RESICO",
            sty("ps2", fontSize=9, textColor=C_MUTED, fontName="Helvetica",
                leading=13, alignment=TA_CENTER),
        ))
        story.append(Spacer(1, 0.25 * cm))
        story.append(HRFlowable(
            width=cw, thickness=2, color=C_NAVY, spaceAfter=3,
        ))
        story.append(HRFlowable(
            width=cw, thickness=0.5, color=colors.HexColor("#B45309"),
            spaceBefore=2, spaceAfter=0,
        ))
        story.append(Spacer(1, 0.3 * cm))

        # Grid de campos estilo constancia — 4 celdas (2×2)
        def _campo(label, value, align=TA_LEFT):
            return [
                Paragraph(label, s_fld_l),
                Paragraph(value, sty(f"fv{label[:3]}",
                                     fontSize=9, textColor=C_DARK,
                                     fontName="Helvetica-Bold", leading=13,
                                     alignment=align)),
            ]

        campos_t = Table([
            [
                _campo("CONTRIBUYENTE", cfg.PHARMACY_NAME),
                _campo("R.F.C.", rfc_val),
                _campo("PERÍODO", per_label, TA_CENTER),
                _campo("EJERCICIO", str(anio), TA_CENTER),
            ],
            [
                _campo("DOMICILIO FISCAL", cfg.PHARMACY_ADDRESS),
                _campo("TELÉFONO", cfg.PHARMACY_PHONE),
                _campo("FECHA INICIO", fecha_inicio.strftime("%d/%m/%Y"), TA_CENTER),
                _campo("FECHA FIN", fecha_fin.strftime("%d/%m/%Y"), TA_CENTER),
            ],
        ], colWidths=[cw * 0.38, cw * 0.22, cw * 0.20, cw * 0.20])
        campos_t.setStyle(TableStyle([
            ("BOX",           (0, 0), (-1, -1), 0.8, C_NAVY),
            ("INNERGRID",     (0, 0), (-1, -1), 0.4, C_LINE),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 8),
            ("VALIGN",        (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND",    (0, 0), (-1, 0),  colors.HexColor("#F8FAFC")),
        ]))
        story.append(campos_t)
        story.append(Spacer(1, 0.45 * cm))

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE 2 — RESUMEN FINANCIERO (6 campos en tabla formal)
        # ═══════════════════════════════════════════════════════════════════
        story.append(Paragraph(
            "I.  RESUMEN FINANCIERO DEL PERÍODO",
            sty("s2t", fontSize=8, textColor=C_NAVY, fontName="Helvetica-Bold",
                leading=11, spaceBefore=0),
        ))
        story.append(HRFlowable(width=cw, thickness=0.5, color=C_NAVY,
                                spaceAfter=4, spaceBefore=2))

        # 3 columnas: concepto | importe | notas
        fin_data = [
            [Paragraph("<b>CONCEPTO</b>", s_th_l),
             Paragraph("<b>IMPORTE</b>", s_th),
             Paragraph("<b>% DEL TOTAL</b>", s_th)],
        ]
        def pct(v): return f"{v/total_ingresos*100:.1f}%" if total_ingresos else "—"
        rows_fin = [
            ("Ingresos Totales del Período",  money(total_ingresos), "100.0%", True),
            ("   · Efectivo",                 money(ef),             pct(ef),  False),
            ("   · Tarjeta bancaria",          money(tj),             pct(tj),  False),
            ("   · Transferencia / depósito",  money(tr),             pct(tr),  False),
            ("Número de transacciones",        str(num_ventas),       f"{dias_con_vta} días hábiles", False),
            ("Ticket promedio por venta",      money(tick_prom),      "—",      False),
        ]
        for concepto, importe, nota, bold in rows_fin:
            fs = s_td_b if bold else s_td
            fr = s_td_br if bold else s_td_r
            fin_data.append([
                Paragraph(concepto, fs),
                Paragraph(importe,  fr),
                Paragraph(nota,     s_td_c),
            ])
        fin_t = Table(fin_data, colWidths=[cw * 0.52, cw * 0.26, cw * 0.22])
        fin_t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  C_NAVY2),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_LIGHT]),
            ("LINEBELOW",      (0, 0), (-1, 0),  1.0, C_NAVY),
            ("LINEBELOW",      (0, 1), (-1, 1),  1.0, C_NAVY2),
            ("BOX",            (0, 0), (-1, -1), 0.6, C_NAVY),
            ("INNERGRID",      (0, 1), (-1, -1), 0.25, C_LINE),
            ("TOPPADDING",     (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
            ("LEFTPADDING",    (0, 0), (-1, -1), 8),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 8),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(fin_t)
        story.append(Spacer(1, 0.5 * cm))

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE 3 — RESUMEN DIARIO
        # ═══════════════════════════════════════════════════════════════════
        story.append(Paragraph(
            "II.  DESGLOSE DIARIO DE INGRESOS",
            sty("s3t", fontSize=8, textColor=C_NAVY, fontName="Helvetica-Bold", leading=11),
        ))
        story.append(HRFlowable(width=cw, thickness=0.5, color=C_NAVY,
                                spaceAfter=4, spaceBefore=2))

        COL_D = [3.5*cm, 1.6*cm, 3.0*cm, 3.0*cm, 3.4*cm, None]
        dia_data = [[
            Paragraph("<b>Fecha</b>",            s_th_l),
            Paragraph("<b>Ops.</b>",              s_th),
            Paragraph("<b>Efectivo</b>",          s_th),
            Paragraph("<b>Tarjeta</b>",           s_th),
            Paragraph("<b>Transferencia</b>",     s_th),
            Paragraph("<b>Total del Día</b>",     s_th),
        ]]
        t_ef = t_tj = t_tr = t_tot = 0.0; t_n = 0
        for d_iso in sorted(por_dia):
            row = por_dia[d_iso]
            tot = row["ef"] + row["tj"] + row["tr"]
            t_ef += row["ef"]; t_tj += row["tj"]; t_tr += row["tr"]
            t_tot += tot;      t_n  += row["n"]
            try:
                d_obj = date.fromisoformat(d_iso)
                d_fmt = d_obj.strftime("%a %d/%m/%Y").capitalize()
            except Exception:
                d_fmt = d_iso
            dia_data.append([
                Paragraph(d_fmt,                                   s_td),
                Paragraph(str(row["n"]),                           s_td_c),
                Paragraph(money(row["ef"]) if row["ef"] else "—",  s_td_r),
                Paragraph(money(row["tj"]) if row["tj"] else "—",  s_td_r),
                Paragraph(money(row["tr"]) if row["tr"] else "—",  s_td_r),
                Paragraph(f"<b>{money(tot)}</b>",
                          sty(f"dt{d_iso}", fontSize=7.5, textColor=C_NAVY2,
                              fontName="Helvetica-Bold", leading=11, alignment=TA_RIGHT)),
            ])
        # Total final row
        dia_data.append([
            Paragraph("<b>TOTAL DEL MES</b>",  s_th_l),
            Paragraph(f"<b>{t_n}</b>",         s_th),
            Paragraph(f"<b>{money(t_ef)}</b>", s_th),
            Paragraph(f"<b>{money(t_tj)}</b>", s_th),
            Paragraph(f"<b>{money(t_tr)}</b>", s_th),
            Paragraph(f"<b>{money(t_tot)}</b>",
                      sty("_ttot", fontSize=8, textColor=colors.HexColor("#FDE68A"),
                          fontName="Helvetica-Bold", leading=12, alignment=TA_RIGHT)),
        ])
        lr = len(dia_data) - 1
        dia_t = Table(dia_data, colWidths=COL_D)
        dia_t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0),      (-1, 0),      C_NAVY2),
            ("BACKGROUND",     (0, lr),     (-1, lr),     C_NAVY),
            ("ROWBACKGROUNDS", (0, 1),      (-1, lr - 1), [C_WHITE, C_LIGHT]),
            ("BOX",            (0, 0),      (-1, -1),     0.6, C_NAVY),
            ("INNERGRID",      (0, 1),      (-1, lr - 1), 0.25, C_LINE),
            ("LINEBELOW",      (0, 0),      (-1, 0),      0.8, C_NAVY),
            ("LINEABOVE",      (0, lr),     (-1, lr),     1.0, C_NAVY),
            ("VALIGN",         (0, 0),      (-1, -1),     "MIDDLE"),
            ("TOPPADDING",     (0, 0),      (-1, -1),     5),
            ("BOTTOMPADDING",  (0, 0),      (-1, -1),     5),
            ("LEFTPADDING",    (0, 0),      (-1, -1),     6),
            ("RIGHTPADDING",   (0, 0),      (-1, -1),     6),
        ]))
        story.append(KeepTogether(dia_t) if len(dia_data) <= 24 else dia_t)
        story.append(Spacer(1, 0.5 * cm))

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE 4 — DETALLE DE TRANSACCIONES
        # ═══════════════════════════════════════════════════════════════════
        story.append(Paragraph(
            f"III.  DETALLE DE TRANSACCIONES  ({num_ventas} registros)",
            sty("s4t", fontSize=8, textColor=C_NAVY, fontName="Helvetica-Bold", leading=11),
        ))
        story.append(HRFlowable(width=cw, thickness=0.5, color=C_NAVY,
                                spaceAfter=4, spaceBefore=2))

        pago_lbl = {"efectivo": "Efectivo", "tarjeta": "Tarjeta", "transferencia": "Transf."}
        v_data = [[
            Paragraph("<b>Folio</b>",   s_th),
            Paragraph("<b>Fecha</b>",   s_th),
            Paragraph("<b>Hora</b>",    s_th),
            Paragraph("<b>Cajero</b>",  s_th_l),
            Paragraph("<b>Importe</b>", s_th),
            Paragraph("<b>Pago</b>",    s_th),
        ]]
        for v in ventas:
            cajero  = (v.usuario.nombre or v.usuario.username) if v.usuario else "—"
            pv      = v.metodo_pago.value if v.metodo_pago else ""
            fecha_s = v.creado_en.strftime("%d/%m/%Y") if v.creado_en else "—"
            hora_s  = v.creado_en.strftime("%H:%M")    if v.creado_en else "—"
            v_data.append([
                Paragraph(str(v.folio or v.id), s_td_b),
                Paragraph(fecha_s,              s_mut),
                Paragraph(hora_s,               s_mut_c),
                Paragraph((cajero or "—")[:22], s_td),
                Paragraph(money(v.total),        s_td_br),
                Paragraph(pago_lbl.get(pv, pv), s_mut_c),
            ])
        v_t = Table(v_data, colWidths=[2.0*cm, 2.8*cm, 1.6*cm, None, 2.8*cm, 2.2*cm])
        v_t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  C_NAVY2),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_LIGHT]),
            ("BOX",            (0, 0), (-1, -1), 0.6, C_NAVY),
            ("INNERGRID",      (0, 1), (-1, -1), 0.25, C_LINE),
            ("LINEBELOW",      (0, 0), (-1, 0),  0.8, C_NAVY),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",     (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
            ("LEFTPADDING",    (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
        ]))
        story.append(v_t)
        story.append(Spacer(1, 0.6 * cm))

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE 5 — CIERRE FORMAL CON FIRMAS
        # ═══════════════════════════════════════════════════════════════════
        story.append(Paragraph(
            "IV.  RESUMEN DE CIERRE Y FIRMAS",
            sty("s5t", fontSize=8, textColor=C_NAVY, fontName="Helvetica-Bold", leading=11),
        ))
        story.append(HRFlowable(width=cw, thickness=0.5, color=C_NAVY,
                                spaceAfter=4, spaceBefore=2))

        # Total box — ancho completo
        total_box = Table([[
            Table([
                [Paragraph("TOTAL DE INGRESOS A DECLARAR AL SAT",
                           sty("tb1", fontSize=7, textColor=C_MUTED,
                               fontName="Helvetica-Bold", leading=10,
                               alignment=TA_CENTER))],
                [Paragraph(money(total_ingresos),
                           sty("tb2", fontSize=20, textColor=C_NAVY,
                               fontName="Helvetica-Bold", leading=24,
                               alignment=TA_CENTER))],
                [Paragraph(f"Período: {per_label}  ·  RFC: {rfc_val}",
                           sty("tb3", fontSize=7, textColor=C_MUTED,
                               fontName="Helvetica", leading=10,
                               alignment=TA_CENTER))],
            ], colWidths=[cw - 0.4 * cm]),
        ]], colWidths=[cw])
        total_box.setStyle(TableStyle([
            ("BOX",           (0, 0), (-1, -1), 1.2, C_NAVY),
            ("LINEBELOW",     (0, 0), (-1, 0),  4,   C_NAVY),
            ("BACKGROUND",    (0, 0), (-1, -1), C_LIGHT),
            ("TOPPADDING",    (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
            ("LEFTPADDING",   (0, 0), (-1, -1), 0),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 0),
        ]))
        story.append(total_box)
        story.append(Spacer(1, 0.9 * cm))

        # 3 líneas de firma simétricas — más altas
        class _SigLineWide(_SigLine):
            def __init__(self, w, label, subtext=""):
                super().__init__(w, label, subtext)
                self.height = 2.4 * cm  # más espacio para firmar

        sw = cw / 3
        sig_row = Table([[
            _SigLineWide(sw, "Firma y Sello del Contador Público",
                         "Cédula Profesional / RFC del Contador"),
            _SigLineWide(sw, "Vo. Bo. del Responsable del Negocio",
                         cfg.PHARMACY_NAME),
            _SigLineWide(sw, "Fecha de Revisión y Entrega",
                         "dd / mm / aaaa"),
        ]], colWidths=[sw, sw, sw])
        sig_row.setStyle(TableStyle([
            ("LEFTPADDING",   (0, 0), (-1, -1), 6),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 6),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("VALIGN",        (0, 0), (-1, -1), "BOTTOM"),
        ]))
        story.append(sig_row)
        story.append(Spacer(1, 0.9 * cm))

        # ═══════════════════════════════════════════════════════════════════
        # BLOQUE QR — estilo constancia SAT / situación fiscal
        # ═══════════════════════════════════════════════════════════════════
        import qrcode as _qrcode
        import io as _io
        from reportlab.platypus import Image as RLImage

        qr_content = (
            f"CONTRIBUYENTE:{cfg.PHARMACY_NAME} "
            f"RFC:{rfc_val} "
            f"REGIMEN:RESICO "
            f"PERIODO:{per_label} "
            f"INGRESOS:{money(total_ingresos)} "
            f"TRANSACCIONES:{num_ventas} "
            f"DOMICILIO:{cfg.PHARMACY_ADDRESS} "
            f"GENERADO:{now_str}"
        )

        qr_img = _qrcode.QRCode(
            version=None,
            error_correction=_qrcode.constants.ERROR_CORRECT_M,
            box_size=10,
            border=2,
        )
        qr_img.add_data(qr_content)
        qr_img.make(fit=True)
        pil_img = qr_img.make_image(fill_color="black", back_color="white")
        qr_buf = _io.BytesIO()
        pil_img.save(qr_buf, format="PNG")
        qr_buf.seek(0)

        qr_size = 2.8 * cm
        qr_rl = RLImage(qr_buf, width=qr_size, height=qr_size)

        # QR izquierda, info derecha
        info_col = cw - qr_size - 0.4 * cm
        qr_block = Table([[
            qr_rl,
            [
                Paragraph(
                    "VERIFICACIÓN DEL DOCUMENTO",
                    sty("qrtit", fontSize=7, textColor=C_NAVY,
                        fontName="Helvetica-Bold", leading=10),
                ),
                Spacer(1, 5),
                Paragraph(
                    f"<b>Contribuyente:</b> {cfg.PHARMACY_NAME}<br/>"
                    f"<b>RFC:</b> {rfc_val}<br/>"
                    f"<b>Régimen:</b> Simplificado de Confianza (RESICO)<br/>"
                    f"<b>Período:</b> {per_label}&nbsp;&nbsp;&nbsp;"
                    f"<b>Total declarado:</b> {money(total_ingresos)}<br/>"
                    f"<b>Domicilio:</b> {cfg.PHARMACY_ADDRESS}<br/>"
                    f"<b>Generado:</b> {now_str}",
                    sty("qrdet", fontSize=7, textColor=C_MUTED,
                        fontName="Helvetica", leading=11),
                ),
                Spacer(1, 5),
                Paragraph(
                    "Escanee el código QR para verificar los datos de esta declaración.",
                    sty("qrhint", fontSize=6, textColor=C_MUTED,
                        fontName="Helvetica", leading=9),
                ),
            ],
        ]], colWidths=[qr_size + 0.4 * cm, info_col])
        qr_block.setStyle(TableStyle([
            ("BOX",           (0, 0), (-1, -1), 0.6, C_NAVY),
            ("LINEBEFORE",    (1, 0), (1, -1),  0.4, C_LINE),
            ("BACKGROUND",    (0, 0), (-1, -1), C_LIGHT),
            ("TOPPADDING",    (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
            ("LEFTPADDING",   (0, 0), (0, -1),  8),
            ("RIGHTPADDING",  (0, 0), (0, -1),  8),
            ("LEFTPADDING",   (1, 0), (1, -1),  10),
            ("RIGHTPADDING",  (1, 0), (1, -1),  10),
            ("ALIGN",         (0, 0), (0, -1),  "CENTER"),
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ]))
        story.append(Spacer(1, 1.8 * cm))
        story.append(KeepTogether(qr_block))
        story.append(Spacer(1, 0.5 * cm))
        story.append(HRFlowable(width=cw, thickness=0.4, color=C_LINE, spaceAfter=5))
        story.append(Paragraph(
            "Este documento ha sido generado electrónicamente por el sistema de Punto de Venta. "
            "Los ingresos consignados corresponden a las ventas realizadas y registradas durante el período indicado "
            "para efectos de declaración mensual bajo el Régimen Simplificado de Confianza (RESICO) conforme al "
            "Título IV, Capítulo II, Sección IV de la Ley del Impuesto Sobre la Renta (LISR). "
            "Este reporte no sustituye al CFDI, ni a la declaración fiscal oficial presentada ante el SAT.",
            s_legal,
        ))

    finally:
        db.close()

    # ── Build ─────────────────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp.close()
    try:
        doc = SimpleDocTemplate(
            tmp.name,
            pagesize=letter,
            rightMargin=1.8*cm, leftMargin=1.8*cm,
            topMargin=1.9*cm, bottomMargin=2.4*cm,
            title=f"Declaración RESICO {per_label} — {cfg.PHARMACY_NAME}",
        )
        doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
        with open(tmp.name, "rb") as f:
            data = f.read()
    finally:
        try: os.unlink(tmp.name)
        except Exception: pass

    return data


@router.get("/cierre-mensual-pdf")
def cierre_mensual_pdf(
    mes:  int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020, le=2100),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    data     = _build_cierre_mensual_pdf_bytes(mes, anio)
    filename = f"Cierre_RESICO_{MESES_ES[mes-1]}_{anio}.pdf"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class _SaveCierrePdfBody(BaseModel):
    path: str


@router.post("/save-cierre-mensual-pdf")
def save_cierre_mensual_pdf(
    body: _SaveCierrePdfBody,
    mes:  int = Query(..., ge=1, le=12),
    anio: int = Query(..., ge=2020, le=2100),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    if not body.path:
        raise HTTPException(status_code=400, detail="path requerido")
    data = _build_cierre_mensual_pdf_bytes(mes, anio)
    with open(body.path, "wb") as f:
        f.write(data)
    return {"ok": True, "path": body.path}


def _build_ventas_pdf_bytes(fecha_inicio: date, fecha_fin: date) -> bytes:  # noqa: C901
    """Generate ventas PDF for given period; return raw bytes."""
    import app.config as cfg
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable,
                                    KeepTogether, PageBreak)

    W, H = letter
    fi, ff = _rango(fecha_inicio, fecha_fin)
    now_str  = datetime.now().strftime("%d/%m/%Y %H:%M")
    per_label = f"{fecha_inicio.strftime('%d/%m/%Y')} — {fecha_fin.strftime('%d/%m/%Y')}"

    # Palette
    C_BLUE  = colors.HexColor("#1d2140")
    C_BLUE_L = colors.HexColor("#EFF6FF")
    C_BLUE_M = colors.HexColor("#BFDBFE")
    C_DARK  = colors.HexColor("#0F172A")
    C_MUTED = colors.HexColor("#64748B")
    C_GREEN = colors.HexColor("#16A34A")
    C_GRN_L = colors.HexColor("#DCFCE7")
    C_AMBER = colors.HexColor("#D97706")
    C_AMB_L = colors.HexColor("#FEF3C7")
    C_PURP  = colors.HexColor("#7C3AED")
    C_PURP_L = colors.HexColor("#F5F3FF")
    C_GRAY  = colors.HexColor("#F1F5F9")
    C_GRAY_B = colors.HexColor("#E2E8F0")
    C_WHITE = colors.white
    C_RED   = colors.HexColor("#EF4444")

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        return ParagraphStyle(name, parent=styles["Normal"], **kw)

    s_body  = sty("_b",  fontSize=8,  textColor=C_DARK,  fontName="Helvetica",      leading=12)
    s_bold  = sty("_bb", fontSize=8,  textColor=C_DARK,  fontName="Helvetica-Bold", leading=12)
    s_hdr   = sty("_h",  fontSize=8,  textColor=C_WHITE, fontName="Helvetica-Bold", leading=12, alignment=TA_CENTER)
    s_ctr   = sty("_c",  fontSize=8,  textColor=C_DARK,  fontName="Helvetica",      leading=12, alignment=TA_CENTER)
    s_right = sty("_r",  fontSize=8,  textColor=C_DARK,  fontName="Helvetica",      leading=12, alignment=TA_RIGHT)
    s_mut   = sty("_m",  fontSize=7,  textColor=C_MUTED, fontName="Helvetica",      leading=11)

    cw = W - 3.0 * cm
    story = []

    def _on_cover(canvas, doc):
        pass

    def _on_page(canvas, doc):
        canvas.saveState()
        canvas.setStrokeColorRGB(0.886, 0.91, 0.941)
        canvas.setLineWidth(0.5)
        canvas.line(1.5 * cm, H - 1.6 * cm, W - 1.5 * cm, H - 1.6 * cm)
        canvas.setFont("Helvetica-Bold", 7)
        canvas.setFillColorRGB(0.114, 0.129, 0.251)
        canvas.drawString(1.5 * cm, H - 1.3 * cm, cfg.PHARMACY_NAME)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColorRGB(0.392, 0.455, 0.545)
        canvas.drawRightString(W - 1.5 * cm, H - 1.3 * cm, f"Reporte de Ventas  ·  {per_label}")
        canvas.line(1.5 * cm, 1.5 * cm, W - 1.5 * cm, 1.5 * cm)
        canvas.drawString(1.5 * cm, 1.2 * cm, now_str)
        canvas.drawCentredString(W / 2, 1.2 * cm, f"{cfg.PHARMACY_ADDRESS}  ·  {cfg.PHARMACY_PHONE}")
        canvas.drawRightString(W - 1.5 * cm, 1.2 * cm, f"Página {doc.page}")
        canvas.restoreState()

    # ── Query data ───────────────────────────────────────────────────────────
    db = get_db_session()
    try:
        ventas = (
            db.query(Venta)
            .options(selectinload(Venta.usuario),
                     selectinload(Venta.cliente))
            .filter(Venta.creado_en >= fi, Venta.creado_en <= ff,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True))
            .order_by(Venta.creado_en)
            .all()
        )

        total_ventas = sum(v.total for v in ventas)
        num_ventas   = len(ventas)
        ef           = sum(v.total for v in ventas if v.metodo_pago.value == "efectivo")
        tj           = sum(v.total for v in ventas if v.metodo_pago.value == "tarjeta")
        tr           = sum(v.total for v in ventas if v.metodo_pago.value == "transferencia")
        tick_prom    = total_ventas / num_ventas if num_ventas else 0.0

        # Top productos
        top = (
            db.query(
                Producto.nombre,
                func.sum(ItemVenta.cantidad).label("qty"),
                func.sum(ItemVenta.subtotal).label("ing"),
            )
            .join(ItemVenta, ItemVenta.producto_id == Producto.id)
            .join(Venta,     Venta.id == ItemVenta.venta_id)
            .filter(Venta.creado_en >= fi, Venta.creado_en <= ff,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True))
            .group_by(Producto.id, Producto.nombre)
            .order_by(func.sum(ItemVenta.cantidad).desc())
            .limit(15)
            .all()
        )

        def money(v): return f"${v:,.2f}"
        def fmtDT(v): return v.strftime("%d/%m/%Y %H:%M") if v else "—"

        # ── Cover / Header ───────────────────────────────────────────────────
        cover_data = [[
            [
                Paragraph(f"<b>REPORTE DE VENTAS</b>",
                          sty("ch", fontSize=20, textColor=C_WHITE, fontName="Helvetica-Bold",
                              leading=24, alignment=TA_CENTER)),
                Spacer(1, 6),
                Paragraph(cfg.PHARMACY_NAME,
                          sty("cn", fontSize=11, textColor=colors.HexColor("#93C5FD"),
                              fontName="Helvetica", alignment=TA_CENTER)),
                Spacer(1, 4),
                Paragraph(f"Período: {per_label}",
                          sty("cp", fontSize=9, textColor=colors.HexColor("#CBD5E1"),
                              fontName="Helvetica", alignment=TA_CENTER)),
            ]
        ]]
        cover_t = Table(cover_data, colWidths=[cw])
        cover_t.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), C_BLUE),
            ("TOPPADDING",    (0, 0), (-1, -1), 20),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 20),
            ("LEFTPADDING",   (0, 0), (-1, -1), 20),
        ]))
        story.append(cover_t)
        story.append(Spacer(1, 0.35 * cm))

        # ── KPI row ──────────────────────────────────────────────────────────
        kpi_items = [
            (money(total_ventas), "Total Ventas",     C_BLUE,  C_BLUE_L),
            (str(num_ventas),     "Transacciones",    C_GREEN, C_GRN_L),
            (money(tick_prom),    "Ticket Promedio",  C_AMBER, C_AMB_L),
            (money(ef),           "Efectivo",         C_GREEN, C_GRN_L),
            (money(tj),           "Tarjeta",          C_BLUE,  C_BLUE_L),
            (money(tr),           "Transferencia",    C_PURP,  C_PURP_L),
        ]
        kpi_cells = []
        kpi_w = []
        for val, lbl, tc, bg in kpi_items:
            cell = Table([[
                [Paragraph(f"<b>{val}</b>",
                           sty(f"kv{lbl[:2]}", fontSize=12, textColor=tc,
                               fontName="Helvetica-Bold", leading=15, alignment=TA_CENTER)),
                 Paragraph(lbl, sty(f"kl{lbl[:2]}", fontSize=7, textColor=C_MUTED,
                                    fontName="Helvetica", alignment=TA_CENTER))]
            ]], colWidths=[cw / 6 - 0.2 * cm])
            cell.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), bg),
                ("BOX",           (0, 0), (-1, -1), 0.5, C_GRAY_B),
                ("TOPPADDING",    (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
                ("LEFTPADDING",   (0, 0), (-1, -1), 4),
                ("RIGHTPADDING",  (0, 0), (-1, -1), 4),
            ]))
            kpi_cells.append(cell)
            kpi_w.append(cw / 6)
        kpi_row = Table([kpi_cells], colWidths=kpi_w)
        kpi_row.setStyle(TableStyle([
            ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",    (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
            ("LEFTPADDING",   (0, 0), (-1, -1), 3),
            ("RIGHTPADDING",  (0, 0), (-1, -1), 3),
        ]))
        story.append(KeepTogether(kpi_row))
        story.append(Spacer(1, 0.35 * cm))

        # ── Top productos ─────────────────────────────────────────────────────
        if top:
            top_data = [[Paragraph("<b>#</b>", s_hdr),
                         Paragraph("<b>Producto</b>", s_hdr),
                         Paragraph("<b>Unidades</b>", s_hdr),
                         Paragraph("<b>Ingreso</b>", s_hdr)]]
            medals = ["#F59E0B", "#94A3B8", "#B45309"]
            for i, r in enumerate(top):
                clr = colors.HexColor(medals[i]) if i < 3 else C_MUTED
                top_data.append([
                    Paragraph(f"<b>{i+1}</b>",
                               sty(f"tm{i}", fontSize=9, textColor=clr,
                                   fontName="Helvetica-Bold", alignment=TA_CENTER)),
                    Paragraph(r.nombre or "—", s_body),
                    Paragraph(str(int(r.qty or 0)),
                               sty(f"tq{i}", fontSize=8, textColor=C_BLUE,
                                   fontName="Helvetica-Bold", alignment=TA_CENTER)),
                    Paragraph(money(r.ing or 0),
                               sty(f"ti{i}", fontSize=8, textColor=C_GREEN,
                                   fontName="Helvetica-Bold", alignment=TA_RIGHT)),
                ])
            top_t = Table(top_data, colWidths=[1.2 * cm, None, 3 * cm, 3.5 * cm])
            top_t.setStyle(TableStyle([
                ("BACKGROUND",     (0, 0), (-1, 0),  C_BLUE),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_GRAY]),
                ("GRID",           (0, 0), (-1, -1), 0.25, C_GRAY_B),
                ("LINEBELOW",      (0, 0), (-1, 0),  1.5, C_BLUE),
                ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING",     (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING",  (0, 0), (-1, -1), 5),
                ("LEFTPADDING",    (0, 0), (-1, -1), 6),
                ("RIGHTPADDING",   (0, 0), (-1, -1), 6),
            ]))
            # Section label
            hdr_top = Table([[Paragraph(
                "<b>🏆  Top Productos Más Vendidos</b>",
                sty("tph", fontSize=9, textColor=C_WHITE, fontName="Helvetica-Bold", leading=13)
            )]], colWidths=[cw])
            hdr_top.setStyle(TableStyle([
                ("BACKGROUND",    (0, 0), (-1, -1), C_BLUE),
                ("TOPPADDING",    (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("LEFTPADDING",   (0, 0), (-1, -1), 10),
            ]))
            story.append(hdr_top)
            story.append(Spacer(1, 2))
            story.append(KeepTogether(top_t))
            story.append(Spacer(1, 0.35 * cm))

        # ── Ventas table ─────────────────────────────────────────────────────
        hdr_v = Table([[Paragraph(
            f"<b>📋  Detalle de Ventas  ·  {num_ventas} registros</b>",
            sty("vh", fontSize=9, textColor=C_WHITE, fontName="Helvetica-Bold", leading=13)
        )]], colWidths=[cw])
        hdr_v.setStyle(TableStyle([
            ("BACKGROUND",    (0, 0), (-1, -1), C_BLUE),
            ("TOPPADDING",    (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ("LEFTPADDING",   (0, 0), (-1, -1), 10),
        ]))
        story.append(hdr_v)
        story.append(Spacer(1, 2))

        pago_icon = {"efectivo": "💵", "tarjeta": "💳", "transferencia": "🏦"}
        v_data = [[
            Paragraph("<b>Folio</b>",   s_hdr),
            Paragraph("<b>Fecha</b>",   s_hdr),
            Paragraph("<b>Cajero</b>",  s_hdr),
            Paragraph("<b>Cliente</b>", s_hdr),
            Paragraph("<b>Total</b>",   s_hdr),
            Paragraph("<b>Pago</b>",    s_hdr),
        ]]
        for v in ventas:
            cajero  = (v.usuario.nombre or v.usuario.username) if v.usuario else "—"
            cliente = v.cliente.nombre if v.cliente else "Público gral."
            pago_val = v.metodo_pago.value if v.metodo_pago else ""
            icon = pago_icon.get(pago_val, "")
            is_ef   = pago_val == "efectivo"
            is_tj   = pago_val == "tarjeta"
            is_tr   = pago_val == "transferencia"
            tc = C_GREEN if is_ef else (C_BLUE if is_tj else C_PURP)
            v_data.append([
                Paragraph(str(v.folio or v.id), sty(f"vf{v.id}", fontSize=7, textColor=C_BLUE,
                                                     fontName="Helvetica-Bold", leading=11)),
                Paragraph(fmtDT(v.creado_en), s_mut),
                Paragraph(cajero or "—",    s_mut),
                Paragraph((cliente or "Público gral.")[:28], s_mut),
                Paragraph(money(v.total),
                           sty(f"vt{v.id}", fontSize=8, textColor=C_DARK,
                               fontName="Helvetica-Bold", alignment=TA_RIGHT, leading=12)),
                Paragraph(f"{icon} {pago_val}",
                           sty(f"vp{v.id}", fontSize=7, textColor=tc,
                               fontName="Helvetica-Bold", leading=11)),
            ])
        v_t = Table(v_data, colWidths=[2 * cm, 3.2 * cm, 3.2 * cm, 4 * cm, 2.8 * cm, 2.8 * cm])
        v_t.setStyle(TableStyle([
            ("BACKGROUND",     (0, 0), (-1, 0),  C_BLUE),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [C_WHITE, C_GRAY]),
            ("GRID",           (0, 0), (-1, -1), 0.25, C_GRAY_B),
            ("LINEBELOW",      (0, 0), (-1, 0),  1.5, C_BLUE),
            ("VALIGN",         (0, 0), (-1, -1), "MIDDLE"),
            ("TOPPADDING",     (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING",  (0, 0), (-1, -1), 4),
            ("LEFTPADDING",    (0, 0), (-1, -1), 5),
            ("RIGHTPADDING",   (0, 0), (-1, -1), 5),
        ]))
        story.append(v_t)

    finally:
        db.close()

    # ── Build PDF ─────────────────────────────────────────────────────────────
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
    tmp.close()
    try:
        doc = SimpleDocTemplate(
            tmp.name,
            pagesize=letter,
            rightMargin=1.5 * cm, leftMargin=1.5 * cm,
            topMargin=2.0 * cm, bottomMargin=2.2 * cm,
            title=f"Reporte Ventas {per_label} — {cfg.PHARMACY_NAME}",
        )
        doc.build(story, onFirstPage=_on_cover, onLaterPages=_on_page)
        with open(tmp.name, "rb") as f:
            data = f.read()
    finally:
        try: os.unlink(tmp.name)
        except Exception: pass

    return data


@router.get("/export-pdf")
def export_pdf_ventas(
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    data = _build_ventas_pdf_bytes(fecha_inicio, fecha_fin)
    filename = f"Ventas_{fecha_inicio}_{fecha_fin}.pdf"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class _SavePdfBody(BaseModel):
    path: str


@router.post("/save-pdf-to-path")
def save_pdf_to_path(
    body: _SavePdfBody,
    fecha_inicio: date = Query(...),
    fecha_fin:    date = Query(...),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    if not body.path:
        raise HTTPException(status_code=400, detail="path requerido")
    data = _build_ventas_pdf_bytes(fecha_inicio, fecha_fin)
    with open(body.path, "wb") as f:
        f.write(data)
    return {"ok": True, "path": body.path}


@router.delete("/cortes/fantasmas")
def eliminar_cortes_fantasma(bg: BackgroundTasks, payload: dict = Depends(get_current_api_user)):
    """Admin: permanently delete closed shifts with 0 ventas and $0 total."""
    _require_admin(payload)
    db = get_db_session()
    try:
        fantasmas = (
            db.query(CortesCaja)
            .filter(
                CortesCaja.cerrado_en != None,
                CortesCaja.num_ventas == 0,
                CortesCaja.total_ventas == 0,
            )
            .all()
        )
        ids = [c.id for c in fantasmas]
        for c in fantasmas:
            db.delete(c)
        db.commit()
        import app.config as _cfg
        if _cfg.TURSO_SYNC:
            from app.database.sync_service import sync_to_turso, delete_ids_from_turso
            # cortes_caja está en _NO_TURSO_DELETE — igual que retiros_caja, el
            # sync normal nunca borra por ausencia, así que sin este delete
            # explícito el corte fantasma borrado reaparecía en el próximo pull.
            # Síncrono (no bg.add_task): si la app cierra justo después de borrar
            # (p.ej. para instalar una actualización), una tarea en background se
            # pierde antes de llegar a Turso y el corte "resucita" en el próximo pull.
            delete_ids_from_turso("cortes_caja", ids)
            bg.add_task(sync_to_turso)
        return {"ok": True, "eliminados": len(ids), "ids": ids}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        db.close()


@router.get("/rentabilidad-cajero")
def rentabilidad_cajero(
    fecha_inicio: date = Query(...),
    fecha_fin: date = Query(...),
    payload: dict = Depends(get_current_api_user),
):
    _require_admin(payload)
    db = get_db_session()
    try:
        fi, ff = _rango(fecha_inicio, fecha_fin)
        from app.database.models import Usuario
        usuarios = db.query(Usuario).filter(Usuario.activo == True).all()
        result = []
        for u in usuarios:
            ventas = (
                db.query(Venta)
                .filter(
                    Venta.usuario_id == u.id,
                    Venta.creado_en >= fi,
                    Venta.creado_en <= ff,
                    Venta.estado == EstadoVenta.completada,
                    Venta.eliminado.is_not(True),
                )
                .all()
            )
            if not ventas:
                continue
            total_v = sum(v.total for v in ventas)
            vids = [v.id for v in ventas]
            cost_rows = (
                db.query(ItemVenta.cantidad, Producto.precio_compra)
                .join(Producto, ItemVenta.producto_id == Producto.id)
                .filter(ItemVenta.venta_id.in_(vids))
                .all()
            )
            total_c = sum(r.cantidad * (r.precio_compra or 0.0) for r in cost_rows)
            result.append({
                "cajero": u.nombre or u.username,
                "rol": u.rol.value if u.rol else "",
                "num_ventas": len(ventas),
                "total_ventas": round(total_v, 2),
                "total_costo": round(total_c, 2),
                "ganancia": round(total_v - total_c, 2),
                "ticket_promedio": round(total_v / len(ventas), 2) if ventas else 0.0,
                "margen_pct": round((total_v - total_c) / total_v * 100, 1) if total_v else 0.0,
            })
        result.sort(key=lambda x: -x["total_ventas"])
        return result
    finally:
        db.close()


@router.get("/inventario-excel")
def exportar_inventario_excel(payload: dict = Depends(get_current_api_user)):
    _require_admin(payload)
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from fastapi.responses import Response
    db = get_db_session()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Inventario"
        headers = ["ID", "Código", "Nombre", "Categoría", "Proveedor", "Stock",
                   "Stock Mín.", "Precio Compra", "Precio Venta", "Valor Inventario",
                   "Presentación", "Activo"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(fill_type="solid", fgColor="1d2140")
        productos = (db.query(Producto).filter(Producto.activo == True)
                     .order_by(Producto.nombre).all())
        for row, p in enumerate(productos, 2):
            ws.cell(row=row, column=1, value=p.id)
            ws.cell(row=row, column=2, value=p.codigo_barras or "")
            ws.cell(row=row, column=3, value=p.nombre)
            ws.cell(row=row, column=4, value=p.categoria.nombre if p.categoria else "")
            ws.cell(row=row, column=5, value=p.proveedor.nombre if p.proveedor else "")
            ws.cell(row=row, column=6, value=p.stock)
            ws.cell(row=row, column=7, value=p.stock_minimo)
            ws.cell(row=row, column=8, value=p.precio_compra)
            ws.cell(row=row, column=9, value=p.precio_venta)
            ws.cell(row=row, column=10, value=round(p.stock * p.precio_compra, 2))
            ws.cell(row=row, column=11, value=p.presentacion or "")
            ws.cell(row=row, column=12, value="Sí" if p.activo else "No")
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)
        import io as _io
        buf = _io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            content=buf.read(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=inventario.xlsx"},
        )
    finally:
        db.close()

